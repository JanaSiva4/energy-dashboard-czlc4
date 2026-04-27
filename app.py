import streamlit as st
import pandas as pd
import requests
import io
import re
import base64
import json
import qrcode
from datetime import datetime, date
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.graphics.shapes import Drawing, Circle, String

# --- KONFIGURACE GOOGLE SHEETS ---
GOOGLE_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbxzGV-vnAWMloGczThHXmch7JmgYDNe2WpPzeDeVvGPgcyeRpCEzi4dQfq7IsZWNLt7wg/exec"
FACILITY_SCRIPT_URL = "https://script.google.com/macros/s/AKfycbxzGV-vnAWMloGczThHXmch7JmgYDNe2WpPzeDeVvGPgcyeRpCEzi4dQfq7IsZWNLt7wg/exec"

# --- FONTY S CESKOU DIAKRITIKOU ---
import os as _os

def _registruj_font():
    _base = _os.path.dirname(_os.path.abspath(__file__))
    _mozne_cesty = [
        (_os.path.join(_base, 'fonts', 'DejaVuSans.ttf'),
         _os.path.join(_base, 'fonts', 'DejaVuSans-Bold.ttf')),
        ('/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
         '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'),
    ]
    for _regular, _bold in _mozne_cesty:
        if _os.path.exists(_regular) and _os.path.exists(_bold):
            try:
                pdfmetrics.registerFont(TTFont('DejaVu', _regular))
                pdfmetrics.registerFont(TTFont('DejaVu-Bold', _bold))
                return 'DejaVu', 'DejaVu-Bold'
            except Exception:
                continue
    return 'Helvetica', 'Helvetica-Bold'

PDF_FONT, PDF_FONT_BOLD = _registruj_font()


def odeslat_do_google_sheets(res, sklad="CZLC4"):
    try:
        obdobi_raw = str(res.get('obdobi', datetime.now().strftime('%Y-%m')))
        try:
            if '-' in obdobi_raw and len(obdobi_raw) == 7:
                rok, mesic = map(int, obdobi_raw.split('-'))
            elif '.' in obdobi_raw:
                casti = obdobi_raw.split('.')
                if len(casti) == 2:
                    mesic, rok = int(casti[0]), int(casti[1])
                elif len(casti) == 3:
                    mesic, rok = int(casti[1]), int(casti[2])
                else:
                    rok, mesic = datetime.now().year, datetime.now().month
            else:
                rok, mesic = datetime.now().year, datetime.now().month
        except Exception:
            rok, mesic = datetime.now().year, datetime.now().month

        def to_f(val):
            if not val or str(val).lower() == 'n/a':
                return 0.0
            try:
                s = str(val).replace('\xa0', '').replace(' ', '')
                s = re.sub(r'[^0-9,.]', '', s)
                if ',' in s and '.' in s:
                    s = s.replace(',', '')
                s = s.replace(',', '.')
                return float(s)
            except Exception:
                return 0.0

        data_row = [
            str(rok), str(mesic).zfill(2),
            to_f(res.get('el_spotreba_kwh', 0)), 0.0,
            to_f(res.get('el_cena_sil_el_bez_dph', 0)),
            to_f(res.get('el_cena_distribuce_bez_dph', 0)),
            to_f(res.get('el_cena_celkem_zaklad_kc', 0)),
            to_f(res.get('fsx_spotreba_kwh', 0)), 0.0,
            to_f(res.get('fsx_cena_bez_dph', 0)),
            to_f(res.get('plyn_spotreba_kwh', 0)), 0.0,
            to_f(res.get('plyn_cena_celkem_zaklad_kc', 0)),
            to_f(res.get('voda_spotreba_m3', 0)), 0.0,
            to_f(res.get('voda_cena_bez_dph', 0)),
        ]
        payload = {"action": "append", "sheet": sklad, "row": data_row}
        requests.post(GOOGLE_SCRIPT_URL, json=payload)
        return True
    except Exception as e:
        st.error(f"Chyba: {e}")
        return False


def odeslat_mcdp_do_sheets(data: dict, sklad: str = "CZLC4") -> bool:
    def yn(val):
        return "ANO" if val else "NE"
    try:
        row = [
            f"MCDP-{sklad}-{datetime.now().strftime('%Y%m%d%H%M')}",
            data.get("datum_vydeje", datetime.now().strftime("%d.%m.%Y")),
            data.get("kvartal", ""), datetime.now().year, sklad,
            data.get("zamestnanec", ""), data.get("email", ""),
            yn(data.get("rucnik")), yn(data.get("mydlo")),
            yn(data.get("ariel")), yn(data.get("krem")), yn(data.get("solvina")),
            yn(all([data.get("rucnik"), data.get("mydlo"), data.get("ariel"),
                    data.get("krem"), data.get("solvina")])),
            data.get("zadal", ""), datetime.now().strftime("%d.%m.%Y %H:%M"),
        ]
        payload = {"action": "append", "sheet": f"MCDP_{sklad}", "row": row}
        r = requests.post(FACILITY_SCRIPT_URL, json=payload, timeout=10)
        return r.status_code == 200
    except Exception as e:
        st.error(f"Chyba odesilani MCDP: {e}")
        return False


def odeslat_oopp_do_sheets(data: dict, sklad: str = "CZLC4") -> bool:
    def stav_exp(exp_str):
        if not exp_str:
            return "—"
        try:
            p = exp_str.split("/")
            exp = datetime(int(p[1]), int(p[0]), 1)
            dnes = datetime.now()
            if exp < dnes:
                return "expirovano"
            if (exp - dnes).days <= 60:
                return "brzy expiruje"
            return "v poradku"
        except Exception:
            return "—"

    try:
        exp = data.get("expirace", "")
        row = [
            f"OOPP-{sklad}-{datetime.now().strftime('%Y%m%d%H%M%S')}",
            datetime.now().strftime("%d.%m.%Y"), sklad,
            data.get("zamestnanec", ""), data.get("email", ""),
            data.get("pomucka", ""), data.get("velikost", ""),
            exp, stav_exp(exp), "",
            "ANO" if data.get("podpis") else "NE",
            data.get("zadal", ""), datetime.now().strftime("%d.%m.%Y %H:%M"),
        ]
        payload = {"action": "append", "sheet": f"OOPP_{sklad}", "row": row}
        r = requests.post(FACILITY_SCRIPT_URL, json=payload, timeout=10)
        return r.status_code == 200
    except Exception as e:
        st.error(f"Chyba odesilani OOPP: {e}")
        return False


# ═══════════════════════════════════════════════════════════════════
# PDF PROTOKOLY — Alza styl (bílý papír, originální logo)
# ═══════════════════════════════════════════════════════════════════

# Originální logo Alza.cz (extrahováno z Alza PDF šablony, 3× upscale)
_ALZA_LOGO_B64 = (
    "/9j/4AAQSkZJRgABAQAAAQABAAD/2wBDAAMCAgICAgMCAgIDAwMDBAYEBAQEBAgGBgUGCQgKCgkI"
    "CQkKDA8MCgsOCwkJDRENDg8QEBEQCgwSExIQEw8QEBD/2wBDAQMDAwQDBAgEBAgQCwkLEBAQEBAQ"
    "EBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBAQEBD/wAARCAJYAlgDASIA"
    "AhEBAxEB/8QAHgABAAEEAwEBAAAAAAAAAAAAAAECAwgJBAYHBQr/xABKEAACAQMCAwYDBQUEBggH"
    "AAAAAQIDBBEFBgcSIQgJEzFBUSIyYRRxgZHRFRYjQlJjcqGxGTNikpPBGCQ1Q1VWgpQlJnN0g/Dx"
    "/8QAGwEBAAIDAQEAAAAAAAAAAAAAAAECAwUGBAf/xAAyEQEAAgIBBAEDAgUDBAMAAAAAAQIDEQQF"
    "EiExQRMiUQYUIzJScbEzQmEVgZGh4fDx/9oADAMBAAIRAxEAPwDamAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAABsAAAAI5kBII5kMoCQAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFKTyVAEb"
    "Ahp+jJKeb6EhhjDHN9BzfQr2g02xhh5THMR3SGGMP2GWMstEyb0hxbCi0MP2JSZO0d0ieET1x0KS"
    "pPoRM6Tr5RhlRCeSSKztEhD80SQ11yTJCQATCQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACH5oIg65wiMMqLcnNfKUm6UgmLePiKXKSxyrJMW"
    "2bTnIHUFtI7olUmhlFDUV1kylzjjCT/BE6T4lcb9mE/qW4tS8sr70JQrelWK+9BGo/KsmLi8pNM6"
    "9uXe+3Nn20rrcWq29GnFOT+JZS+4+Fw040cN+Kta+o7C1iN9OweLjlXyvOPctbFbW5jwp9aszrb0"
    "Ah+Rb53LrDqs4Zcb6GOI0v7E8jPXBSVeQmNpSACQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAEEgAAABGUG0vNj2JBTzr06kOcvSDZEoiVZRKLmujwUzqw"
    "gnKrPw1/tNYPmXu49vWUHK73DZ26j5udeKx+bJrjm3qETesfL6OfD+aY8Wm+imkeQ747U3CrYcJu"
    "/wBbo3nh5z4NRM8F3X3pnBnR51LWz025rVeqi/PqevHwc151FXmvnpEb2za5akV8Ki/vZxr6+stO"
    "ou41O7p29NeblPCNVvEbvNN96pGpDYdSdtz/ACZ6YR4PuLtf9pDdamtZ3HUlQnnpGWOhtafp3kz5"
    "tLw36xh9RDcrq/HHg/ojlHU996dSlDzi6/U8W4mdv7hfsihUnoN1Q1eUF0UJ5z+RqA1vXdb3VXdx"
    "q2rXbqN5lioz59C3VF/wrmvLH9csmwwfp+tf9Ty1+bq9rf6fhsI17vb6t3GrY6ZsaVs3mKq5b/zP"
    "AeIPbi4x7rrVKmga5c2Uar6KM2uUx6mpzXxqP+6hTlWp9KTj18+htcPRuLTx2tfk6jybeYl2rVeL"
    "/GXXXOGv74u72nUi8xlWbxky37q7eFLa+va3pdxWderqs8Scp5w8mDrznOUs+Z692T94XGyOOm3N"
    "NspyVPUbiEar9OrMXUun0pintjxpl4XNva8bny3owk4RWI5cssu5yWk3J0vDaacM5LrWDgLTp2OL"
    "zAvMqyiEiH5kQtZWCE1jzJCQENpBPIEgAAAAAIyhlASAAAIyhle49iQAABGUhlASCE8kgAAAAAAA"
    "AACMr3AkAAAAAAIfQCQRn/EZ64AkAAAAAAAAAAAAAAAAAAAABSvMJ9WR7giJ74RWEyfv6FKan1T6"
    "oqfVYLfM4SUYx6P1JivjSJnUoc4TzJS6wOqbg4qbH2y5x13WLehKmviTksnZLulG5oVbSjU5KlSM"
    "lzL0NLXbU1TiPt7j1qWgVtw3FOxUm4rnaTXse7p3CjlZO2ZeTl8mcNdw2Q7p7d/Z82x4kKu4YzqQ"
    "6YWPM8I4id51YW0Kj4e29K7xnl545NaN3Sp39Tn1GnKrL1k5PqWo0KFB8tnB0l951uH9PYI9+XO5"
    "eq5J+WVe6+8n427hlUoTsIW1OXRciweJbx438Ud/SnU1Hcl7aU5PLVOpJHRqk6sIrmqObZFatTsa"
    "Pi3dw1Fr5TZ04OHjx4h4rcnLlnxK6r7cVbMZa9cXcf5vFm3/AJlp0dPb57mlRnW9cljSaO5dz3X2"
    "PaWlVrqcnjEKbeT3jhp2GeL/ABD8Ktquk3OnwqdXKUGuhS/KxYfOoXrhz39vCal5YU/OlBJf0lie"
    "uWSfLGNV/RGxbZfdP0mqdbXtw9H1kpHq2k92Pw205Lx7qlXa94Hjv+ocEfbt6qdHye5hqQ/a9lHr"
    "4Nb8if2xRl1jQqL8DcbHu6+FPJyytaH38har93LwunTcaUKMW/8AYME9fwb9rz0jL+GnqOrW03yz"
    "jJfeciFelPrSmuvllm0vXu6w2LqzlK21yFu3nGIP9DxDiT3Wu4dAhWr7Z1ad2oRcqaim+Yz4uu4L"
    "zrak9Jy1jemFFJU6lTkuZcra6YPt8P8AXI7W4kaPuCq/4VlXhLnfp1G/+FfErhXcVLbeu3biytac"
    "sRuZ02k1951mvc07zQK9e3bfI1yz9+psY5NOTHh4f2t8dn6A+DO86G+OHOmbnpVOenWopKXvhI72"
    "nmKZjF3e2u09W7Nei2NWsp17dSUsvLw0v0MnF0WPbofOOZj+nyb7/Ls+Lb+FWEglLIzjpg8t43Hh"
    "6f8Acgq64wOmejAmd+CI0pw0Th4xgnHu8kN+giCbJSwSUEptEzXaO4WUw11Jyg2VrWYTuFIHN9EM"
    "+mEW7k90JS6h5CaJ8ytqzZG1t8yXQhuMl82MBScJ8r6pnR+NO8pbB4Z67uegv4tnbznDr1yjJx8V"
    "rTFI9ypmyxjrMy7ynJr4KkQ4z5ubxI4NQFj3lfEKcrmCtKzjCtKKeX5ZLz7yfflOlKUadWcvbLNp"
    "PRs0+WvjqmONw26zm+bHiRK22kupqk4d94LxH3fv3SNvR06s43lSMX5+rNo21b681DRravqFB060"
    "6UZST98Hn5XFtxY3Zm4nJjLaZfaj5EkLHkhk1+tvduEgjK8xn0GkbhIIyMonRsePJkNk9GUtfQb0"
    "mIhKzgglPAaFbRKt4n4IsNdckY9iX5EzGysynKGclJMSNLJefREYZURlZJ0janAw/InK9icobTtI"
    "AAAAAAAAAAAAAAAAAAAAC3F+ZDbyIZz1KmlkikdsaRS20xWV1KG2pfMsFXkse/Q6Jv8A4zcOOGVW"
    "NrvTWqdnOceZKWG2vzFp8eFu3cu6KnGNbxYRbeH1NYnee8OJaZqseIFa0lFXM8RqY6Z+8yD4jd4v"
    "wj2xGdHb2owuZRyk/cwo7WfbPq9o7bVHadrZU4W1Cp4imo4bNr0PP28iJ/7Nb1HBvHLF2m67p0p1"
    "HlOKaRclKFVqFNfF7Ec1Ojb0stupTio4KbqrR020eqVX8S68p9EvaMVe6ZcZXeS3bELd3qFHT4qF"
    "Rc9WXyQ9Wz3Xs29iriH2gNYoa3qdrVsNGjNSlGsnHmj+J2vsLdj3UePG6qe/9/W1Sjty1kqtvldK"
    "jTykbhNA2zpG29Pt9K0XT6VlQtIKnDwoqPMkvXBzHUusxT7Ke2+4XT59vIeEvZB4P8LtPtaenaFR"
    "nqFGMXOrJJ5kj2u1p0rSMbS3taVOEFhKCwsF9eKn5Z+pLcc4fmcln5OXNP3Tt0WLDTFHiDGVhfD9"
    "zIlTglmUpP26jLfwvo/fBYubi2tsO6vKFP2UpJGKlLT7Wm8R6X+VzWItr8RGnGKxByz95TRqwrLn"
    "pXFKcX/Q8lzGItx8xeJhNbdynwk+tScs/wB4luSS5Ip4/qZCxVabeHEqXM5vK6CltpmJdA4rcGNi"
    "cZtDq6FvrSKNelKDjGSS5k35PJpZ7U/Zy1/s5b6u7KtBvbt/Wf2KPoo56G+OpDmxH6p5MRu8X4R0"
    "t/cMZbi8BSnotOU3LHtlm76PzZpmjHM+J/8ATW8vjx2Tb5eOd11xGranfXeyIVW7e1ouahny6Gx2"
    "DTnNezNL/dnb9p7L4uXte7qJU7yTt1n69Dc5Qca9KFxB9KyU19zRTrVJrn7/AIlXp+TuiafMOQuh"
    "S+jIa5sJegkuZ8vsaeJ8trCV0fUqXv7kRXl9A37CIRaVRGUH0RZclhqD6iZ0Vrte5k/Jjo+mSxFO"
    "DznOfMlSpKWeZ5XoViZlMxEL3TPmGl5llNTeefla90cW91fT9OhKWo39vTgv6pJMvWL2nUQxzakR"
    "uZc5pY8x0z5o6fX4r8P7ao6NTXrZSj5rnRbp8XuHVWooR1+25vL5l+pk/bZffbKv1Mf5d1wvcl9E"
    "fK03X9K1uKq6XqNvUj9Jps+hKTXnJST9kY71vSdTC1bUt5iUOUc4fV5MV+8C3zV2pwyqaXTq8kdS"
    "pSjJe/oZSS5kk6Sz8Sz+Zrv73LcNbTdL2tp9vUaddtTS/vM9/TI7uVWJeTqkfwJ01x0XVp061Ogo"
    "x5qrlLoV0nTm3CjBKXq8ETio04tPEpJN/kUx8SCbpr0eWfTq4adkS4X6lu57n2N9KvNb41aLUlSh"
    "OFvXhn4fqbu1JQqUqKjjEPQ1Od2dtNa5vC41WNNzlazTy/Q2u3V1aWbhXvLmjS+HHxySOB6/q+eK"
    "V/Dr+mx2Ye6XOxjqx0/M4X7W01pP9oW2H/aL9R+1dN/8Qten9ov1NFFL/htfqU/Lm9PIdPM4X7V0"
    "3/xC16/2i/UftPTo9P2hbdf7RfqOy/4O/H+XN6fmR0xjJwaWpWWZSeoWzX99FUdT03PM9Qtuv9ov"
    "1E0vHwmL0n5cvrknOfMt06tKt8dOcZR908olz6NpZx5YKxMWNTXyuYRS2/VlqNWpNYjKMX/tI4Wp"
    "65pOjpVNU1C3pfSUkmJwzadQictYjcvpJr3Qb9MnTbnivsC2l/G1y2TX+2imjxe4fXE1Clrts2/L"
    "4kZa8TL77ZU/c449y7pleRK6ny9P13RtWkp6dqNvUbXRKaeTmqcqc23LmX0RSaWrOpZPq1t525GV"
    "5Z6lOOuC3hOXi0nlvzReX18yu9J13IUfcnl+pDbGRPlMRpUCnl5csKWcphKoEcqKXUhnDYFRJRyr"
    "5slLnUcvhXQC6AAAAAAAAAAAAAoisMq6LzGCl59SmS+vSKxofWUPo/8Aka4e9R4e67eVLPfdneVK"
    "dpaUYxnCMmk2kbHG2uvsY2d4Dt+hrXZr12u6HiXNGMXDC69U8l8cRaPK1Las0hW9C0u6H26vUdWc"
    "m3yOeclzT5U6dxzUYOn16wycW1s/Ds6ShGVOrScuZP7zmOVNwjWp9Kz6SLcT+DyI1+Tmavjl9l1K"
    "SmrhQ50llo7Bwe4bT41cU9J23c3P2fT6teCrZeFjJ1y3q0acEkvmXU+lt3cut7Nvv2xtuq6FzB80"
    "ZeTyfTss15OGIr+HCYprx8s2lvo4T8PtucJtm6fsPQ3R+zW1LpOLS5md55KuMLlx6dTRRadtPtF2"
    "PL4mv1JwguWCyztOm94d2ktNgvtOuSjBeTk8dDks/Qc1p3Ftuhw9Ux1jWm6+Ea0c/K/xKZqomnKM"
    "VH+Z58kab7PvQOMFk07zWI3Ml5wjLLOyaf3qXEa8069sJ7S1C7rXNKVOlKnTb6tYPF/0fNjnUzD1"
    "/u65I3ESyh7Zvb00vgNU/c/attDU9UuaeHUpzy6bZrp3J2tOPm77+erU9xXdtTlJyVHxGsHSr3an"
    "G3i3u663bq229TdS5qudNVacnhN5WMnatN7K3HjVKjnDS69ONT5U4NYOg4nB42GkRM+Wsz5c1p8O"
    "88E+3jxg2HvPT6G6ry5vrC6rRpy55tqKbxk3KbE3lp2+dr6fuXS68alO5oxqTSeeVtdUaRo9hXtC"
    "alqNjmlKFONRNtxfTqbe+y1w81rhtwvs9u69VlO5jCPM5P6Go6tiw1jdHv4f1PG3rvJzNVYywn1L"
    "jln4YlEU4U1Tx6EpcqxjyOfjt3qGynu0LHk38R5B2r2qXALdkZtPmtJ/5M9dilUmqi6YZ4N219U/"
    "Z/AvcFJ1FHxbeaX16Hr4NNciunn5Nv4U7aW+Cu5rjaW+dJr28nFVtUjGTTx0cz9BW2bhXW2tFuoS"
    "UlUtKMs596aPzs7XnClcWeoS6Str3xc+2JZN63ZP39+//CPTr511VdpShRynnGEbrrmKe2l2n6Rk"
    "i2a8PZ+aMct9MhP1SLdNKpTi37FUZLm5MeRzdodBPidLiWEUtYKsoNZ8iKz+UWgl8rOKoyqyb+Ro"
    "5T6LJZqSzHxEuqeCsx3Tpetu2EuThinFZn6L3PM+L3HnYXB7Ta15ruq28b+nHmVq6nxPocPtFcct"
    "C4J7FvNUv9RpUdUqUZfY4ykk3P0NLfFrjFuvjLua53FujU606sqrwlP4eX0N103ps8qd2/l/y1HP"
    "5v0Y1X2ys4v95jr27a1Wx2dps9OpUW4RqQk/i+pjfuDtM8a9yV5Tluy6VNt5XivoeYOTlSVC1p5f"
    "rLlLlrZ65cVIWWi6fU1OrU6SjQi5OP34OvxdO4/HrG4c1fk8jJbcS+/d8R+JN9XdR7tuFPzf8Vll"
    "cSeJMKi8Pc1xmPr4jPu3fAniLZaDHcEts3zUlzOHhvKR0R07mhcO2ubaVvWh0cJrDyeqtOPaNV8s"
    "M588e3oeh9objTtrFS03ldcsWvh8VmVPZx7xXWtK1Oz0DfEZ3v2yoqLqVJP4W3jJgipTpVcVFzJ+"
    "gqSdBzv7OXJVto+JDHo0eXkdNwZaz4ZsPLzRaJiX6JND1rTtc0uhq2m14VqVzSjVzF5xlZwapu86"
    "3zQ3Ru600qNRTlp83FRz5GU/dm7917ffBG6udwXEqtW0rRowlJ/y4f6Gt7tO7jrbg4+bmtLyo507"
    "S4mo5+85zo/GiOZas/Dd8/kTfjRZ5ek3bxqzfVpLHsU1HK3gnTfOpdCpvxqrcelNLoU1X0o06aac"
    "qqjlnaTeYiYczGvDZr3VO3IWO39a1KvbqNStFOEmv9pHmHb37Qe+9G4p3Oz9rbhq2tO3m+kKjRlP"
    "2IdufuRwVW46tB5q2/icqXWWFk1r9rLV9a3dx01TVrPb13JOpJZ8NvPU5PDiryuo3758OiyZLY+F"
    "Xt9vlLtLcZo04U3vG6zFYf8AFY/6S/Gb/wA4XX/FZ5vLTdxef7sXf1fhsLStzTi5w2neygvmkqUs"
    "R+838cfjemn+ryHpH/SX4zYf/wA4XX/FYp9pTjVGk5S3jdN+n8VnmDhOL5atJ05+sWuqIxVc1B+R"
    "kjhYZ86Y/wBzljxt6eu01xrqLkjvC6Wf7VnL0ztE8abncOi6a95XT+03MISj4r65Z5Op+BNLC+87"
    "zwQ2bqu9+JmiUrC0lcq3u4TkoLPLiXqYs3DwRjnx8MuLkZZvHlvG4S2eq2fD7SXqtd17y4toVJOT"
    "69UdU40dpPh3wT0mtc6trFvW1GnF4tPE+Jteh1ztEcfNJ4CcILa5epUIa3Gzp06NvzrnyoJPp95p"
    "w4j8SNc4u7kr7n3JqNeVSvOUqcOZ4XX2OR6d0qeVbumPG3RcvqP0aREe2VnFzvKN3b2r1bXatlU0"
    "inFuMJRk109zHrcHaS4xa9OSut5XMuby/ivoecTdfHLUcXFLzS8i/pmlalrNVWujaNX1GpJ4zRi5"
    "Y/I6uvS8OCsb1Dnp5+TLbw+xdcQ+Kc0ripui5mn1/wBaymlxN4lYUqe5riEl/as+xf8ABTijpmk/"
    "ti425efZVHmknTfRHSJVVc5tVQdCtTeJKSwZsOLBP2xMMeTJljzL0rQO0rxl25Wpqjuq55YPPN4j"
    "Mz+yz3iM9U1ez2Nv2nl3Eo0/tdSXlnpnJro8VQxGolKK6PoWb6rd6fRep6bWdKtTlFwcejXU8vP6"
    "PjmszEPTw+oX7oiZfoq03ULDULale6VdQubWquaNSEspn0H5ZMb+wfu643R2edGlf3Eqt7Ri41HJ"
    "5b6LBkfH5cNnA8jHOO01n4ddgvF6xMCWRgLOSW8eZgrbfiWaUQfMupEmk+hUsY6FDypZx8Pq8+Rc"
    "VSTaynjBTHln1ceqPj3+7Nr6e5fbN0WFB0+slK4hlfhk+BR408Nri4dtDd1hGUXyturHzA7zzL5c"
    "BvlXRHyrDce29Tcf2fuKyuJT+WNO5g2/w8z6cZyy1KLSXqwLgKZSjCLlOail5tvB8me5tsQqSpT3"
    "Rp0Zx+aDvKSa+9ZA+wD4/wC8+2P/ADRp3/vKf6j959sYz+9GnY/+8p/qB9gHx/3m2x/5o07/AN5T"
    "/UmO5dsynGEdz6c5SeFFXlPL/wAQPrgiMoyipRkmmspp+ZIAAACMZJBGokUuKwzqPFLZttv3Ymp7"
    "WrQU1cUXiLXrh4O4Frl5HUa83Ff8yfSIfnV4o6W9C4ubh2p4fhQ02vUjhfRnXY04Rqc+eh7p20Nj"
    "1Nn8btw6xVpOD1KvOSeMZyzwt0pfZ0s4fnktWdWiVpjdZiX0rJ06lZUZPq10OVGFSdSVG9xTivka"
    "9Th6KvEk6DWXjPiex9Tae1N2cTdz0tmbRsKt/K4qKnOtCDfhJvzyd/03l1vxu+3w4/ncS31tQ+Te"
    "6vZWtaGnULercXVR4pqmubqe48JuxLx/4vqhqV3oVS30Ws0/Eaw+VmfPZR7vPY3CnS6O4eIdhQ13"
    "VK0FUVOvDm8OTMxtO07TtEtYadpFGlaW8V8FGEMKK+hqeZ13t3TF/wCfhsOP0yIrFp/+WC3C/uoe"
    "D2kUKGp7lv7m4v8AClOn80U/XzZkzsjstcHNiKEdP2lYXM6a6OrRi2etfZ4SfO4rPq8FUVHmajBJ"
    "r1Oey8zJknumW2pxo7fL5FHae1KEo+FtewpOPRKNvBHPjpulUsKGk28fbFGJyYSdVOXKsroWLm5p"
    "2NOVe/nCnSX87ljB55tfJ6lm7a4/MwqVtZcyj9hprHk1BHJSSXwrGPJHw7Xd+1ry4+y2OsW9xXTw"
    "4QqZaPtxllJ+jE1tEfcRetv5QqS6dSly+pKZTt15X1KmtiMOnujDfvKt0x0PhnTsJVeX7ZTkks+Z"
    "mRUTkl9Gma2O933NTpW+1tEt6ycpN+JFP6tmz6RHdyqvD1GdceYhro02k6OlXU/JtNo2q91Zuulc"
    "8KL3b91c81ZV+aKk+uOv6mrRQ5YfZI+UorJmH3bu+K+m8YLXZFGq1SuHmUc9DrerYa34ton4hzXT"
    "MkxyNx+W26Ccafhf0omGYwwl8RFSTVbp7LJVFYnz83Rnz6In5dlvcqsPOSqOCht56ExyWmu1e/c6"
    "S210KU22010Km/oW5SkppY6e5WKzHpkjyw87ZPZS3Z2itQtK9reVaVGyfw04t4kYqa93b/FS2WNH"
    "sJza6dI5NsmratY6HZ1dT1e9o2ljQi5Va1SWFBfUxM4994nw74Upx2fVttztPlk6M20pfgbrg87l"
    "Y/GKNtPzONitO7ywW1Hu+u1ja3UKWh6DSlRm+WUpvGEzOfsadi7TOAujz1zf0bC91u+ipVKdblmq"
    "UvpnoYhcQu8r4tb7m/3YrVNDoyXSMFy4PDdX7RXaE1y5qSfEi8k6jbSjVawbuY53Pp2XmIifx7a3"
    "v43Enx5/+/LeJe3PD2FGrZ3lppHhtOM4ctPGMGoTt56NsDROLcbjY1SkoVXmpCk1yp5+h4/LjLxr"
    "pN0bje95VqP5s1X1R1vUtT1LXrl6prt5K5uF5uTy8mXgdOycK/dve2Lk8mvJrrw481KlU8aazzvy"
    "Kb2HgWFzWz1nSaSJXiXkuaTwo+SJof8AWr6jp9ZfDWlyI32+ys2lpqxNckabO+7W1T91eytuHcl4"
    "vDVKdScW/pGWP8zXVxTvVrnFbcOtppq8rTkn79TYvtTS5cKewTq14oeD9opuXtlNGsKpf1dUrPU3"
    "/wB7KTb/ABNF0ukXyZMsf1Nxzss1wRREcOn4Pk16nL0y1q6hfWVnThmbuYeX3nHajOScH1O88CtH"
    "nr3FbTdI8LxM1ovGM+pust4x4LzLVcbFOTJWG67gTotLT+Ce2rKNFNzsKbnFrzb8ybzgBw61XUJ6"
    "vqGgWsrmq25fw0dq2Np8tI2XpNg48vgW8Fj2PszXLUVTKal0wfM5z3+pM1ny7v6NK44iYeby7PXD"
    "CEHS/dy05X1b8NHgHah3zwL7P+yL7QdP0XTa+r6hTlGnFRi5Rk0eldqvtT7Y4AbWuEq1K41yvTao"
    "2ql8XVeeDTbxB31uHihuq63dua/qXCvakpU7eU2/By/Y3/SuHyM0xltM6/y1HP5GHHE44fI1HULj"
    "WNUudUuKMaUqtRyjCPkky1FTVNuqsIiMJQqPxJ9PJM5+39B17deu2+39sWFTUNTuKigrWEMvD9Tr"
    "/r/Qx/d8ObnF9S/2ruz9ra3vncVptTQrOpXvb6cYUXCOV1fqbfOx72QdG4E7fpbh120Vbc13Tfjw"
    "nFNQ+7Pr5nz+xn2OdA4J7ct917ns6d7uDUKUK6jUp5laSfXC+vkZXxfTmqNOfXD9kcP1Tq1+Vacd"
    "J+3/AC6vgdPrgrF7e2C3a67HG9+OPEOjurTrys7amulrl8n5GO+sd2/xahKrPT9Onl5cUo9EbXdd"
    "3Do+1dPq6xuTVaFhYUlmVerLEUYgcb+8r2Rw1vpWG1LWhuBPMY1qcm1n8GW6fz+ZSNYKxKvLw4bf"
    "6s6YZS7u3tTVL6laUNIoxsq0uWpUm8csfc2DdlrsjbQ7Pu1oVtzU9PvdVqxU6sqqjJQl6+ZgfxE7"
    "x7jNvi5qw21f1dJo1ekYQXKonjF92gu0PfVJ0rziDe1nVecKq+hsb05vUPtyzER/w8NL8bi/dTy3"
    "h6pLhvqunV9L1C20n7LcU5QqfDTWFj6GlvtdaBs/bPHjUtJ2JUpz01SlJOn8uTp0+L3GRvwK++r1"
    "c3zJ1WdXutR1DUdRnX1OtK5uZdZVpPLf4mx6f0evGv3xaZeXkc6OXHbaIWISzPqslq+k6lpV9k10"
    "ORNxpv6sqoWrvKsbGCzKvKKS/E3Fp1SdtVWNW8NwPdzaLVsuB1lqNRvkqrCTMsMYbfueGdjDRpaJ"
    "2e9BsZ0+SahJvpj0R7nFYj1fofL+XbfIvP8Ay77ixrBRD5vQpU3LpLzXmSsqTbfQplKGXUckox6y"
    "b9sHmt7eq3pY1HVLLSrKtqeo3NOhZ0IOU5zeEka8O1d3ltbamqXGy+DlGjfyxKjc18ZdP0bT9C13"
    "kfasudIg+EuyNW8GVzDlrXFKfyv1XQ1mWtrcUpylKrKtdVpN1rhyy5NhMQ7ju3i9xC3lrNbUbnd+"
    "oUKlVucqcaskln0Pgx1rdsm663neqS9qzzn8y5tzbe4d4a9Z7U2XpNTU725qqnWcI8zgm+rMkeK/"
    "d7714VcOqPEaNxVuak6Sq1rbDfI2s4CzxLh/xz4l8O9Xp67Zbrv7u4tpqVGjOpJqWDZ/2NO3zZcc"
    "Zrae/vCsdbpR5UsY52v8zUVGhc8qr1qEqdeDeYY6xaOxcOtyajtTfmha5pDqW9x9qh41RPHNHmA3"
    "y8fd0VNqcJ9X120rqnOFHEJN46NPqaJd07s37d731fXbbiBqEaVzVnJwVxLEevp1Nz/EfZd32nOz"
    "na6DoWsuyuL21pKdSMuvMoYZhJZd0VuWVrOd3u+alDLk3L5wrEsK1vnf9ZclvxA1FxXn/wBYl+pS"
    "988Q/knv3U/D9Wq0v1Myd691bZbB2bd7x1ziBGzpW0ObEp4y8dEYS6jY0NDvbvTaFyrulRm4RqZz"
    "zLPmGSJh9H9++I6ait+am4fyvxpfqcnbe+OIr4hbVpy33qUo1L+mpRdeWGuZfU6xCcbWnK+vKjhb"
    "w83/AEmQHZA7KG8u0tvzT9cr0K2l7b0esrmjqDg1GtyvOMhEzDd3tic57X0OdSblKVlbOTfq/CWW"
    "fYODpWnLTdKsdMU+ZWNClQUv6uSCjk5wYwAAAAALM5OEvE9HiJeLc05NJrplMIhqs73zaNLQdS29"
    "uKwtv+0JZm4ry64f+RgPXny06cl15opNG4rvIuGst6cMqet/Z1UjpFOU22vl82abo16Ls727zlUW"
    "0vpgmPbJUoftXWdatNo7ahKpfXdRQfL54Zuh7D3ZO0Hgrsaz3JqljTq65qdGFWpKccuDf3mvXu0+"
    "EFDiFxgo7v1C2de3tav8yylg3bUaVK0pxtaMcU6cUox9Ej1zyrY8f0aT7ebJxq5LxaYSuaC55+a9"
    "M9CnwqVeary6NF3lWed+pHwqLj5Jmvr3Xme5m3FdVJybxy+S8yxe3tnZW0728u4WtGinKc5ySWEs"
    "vzLzy8QgujXmYDd532htQ2vs2nsLZF/OjqdxUcbmUHhxi+j8j2cXjzyskYqsXIzRgp3Lvae7y7am"
    "xr652ZsJSudSptx+003zRUvvRhJuztxdqLesqkbncM4WGXyxTxhHithp1rShG41Sp9sq3a8SrWq9"
    "ZKT+rI1W8hY2dSnB4pZSjh9WdxxujY8GPutDmcvUbZb9tfLMXu6dy8XeI/Ge6qa3rVavZ0U5yUpt"
    "rp5m3eGY01DPWHR/ka8e634RavoGlVeI1zRlGz1CnKMJNeeUbDsYcpe7OR6patc/bHqG84ETandJ"
    "5laSKUg31wa7fc997aHL4pr2iaWO8J3rcbz4tVtLq3HiLTK0oqOfLDNzGr6jQ0ayutQuJqNOnRc8"
    "v6Jmgrj5rT3Bx73XewqudN3VTl9vM6D9P4ZnJN59NN1bJqkRDqT+Dlre6SPW+yFuyOxu0Bpm4p1O"
    "SKnGOWeRTmqlNR9jkafqdTQb2y1S3m41Y3MFle2TreVj78U1lzXDv2ZNv0N6PqVHW9JttVoSThdU"
    "4yi/ozmwlFvwcdV6nnvAHVI6zwZ2texqKc6llTc3n1PRYxSeEuvufNc1ezJNfw7fj276RZV8K6YG"
    "fYjHXCH3mKszMM+oTlYyz5mt63YaBp1xq2qXVO3tLem5zqzaUVg+hPGVH+roYS95Jxjntnh5ccM7"
    "W8la1tUpxfiweJJY9/xM/FxzmyRWYeXk5fpUmWM3bE7bW6eIOt6nw32RfzpaLGUqVzUpywpx8nho"
    "w6srOz05To2lWtU5nzydV5+Iq06hKnYRjVlz1Ytt1X80/vLkeeeFKKS9z6Rw+NhxY4iauO5nNyWt"
    "MRKupVqfZ3KpKEY++MZ+h6Hwf7PXETjhfR0zauj3FjKbShdVINQkvfJPZ64Ww4z8WdO4e3VVU7ep"
    "VjUlLP8Agbv9jbG2/wAPdsaftnRNHoUXp1vCiqtOjFSljzfNjLNd1fqkcHVMUeZevpvTp5Ud+SWt"
    "2z7r3iBou37m/wBc3BazubejKrOSqJrCWcf4GF+5dMoaHuC+0hzVR2VV0puPk2ng3V9rriRacOeF"
    "N9dXmquhUvKUqUUpYk8o0h3dxO61XUtQnUlUV3WlUUpeqbHRubm5FZvn8/hTqOLHxr9uNRKFStPN"
    "u+WKOZo9KNXd227Tkc3XvacJY+skcSGZLEXjJ6J2dNq1N5cWNBsqdLxXbXsJNYzj4jaZr/bb+zwY"
    "qTNos2I9tfVbDZfZM0zYNu1TqajZUpJL+7n/AJmqG1mrbTreza6pvLNgXeibknYX22Np+Jy8lnCP"
    "L+CMA5xpyhGi+kkzX9Br3cW1p9zMvX1S+8kV+NQmVOVu+ZPK80ZJd37t+jrHaH06rqNDxKSaabXQ"
    "xtrTlToTpx6uMcozp7snbFS53Kty3dB89KWIywZ+pz28a0f8HTY/iRP4bRpRdN+FTXwRgsRR4v2m"
    "+0htPs+bKuNX1G6hX1CtFwt7aEk5xnjo2j6PaH4/7U4EbSutZ1u+pLU50ZfYaGes5ehpg4u8Yd08"
    "aN43e8Nz3lScalV8tpJtwjH0aRx/TOnW5Fu+0eP8t91DnxjiaR7cDihxP3Zxj3bc753ZfVKsnNyt"
    "abeVGGeiwdXk41m69NYnLq4lqvNTkp2Sfhrzj6I+jpmi6nuPWLHbe2KEq2t6hOMLaml0cn7ncVyR"
    "x6+IcvNZzz5nyo07TtR1+/tNA0awqX97fzVGNOim5Um3jLx5G23sU9i7SOB2h2289629K83dcKNW"
    "lVbz4MZRzj7+v+BxexP2JNM4JadS37vu2pXO8L+k/Gt6sVKFDPtnyf4GYCpV41HLxoOk/R/y/RHI"
    "dX6rPJt9PH6/y6XpnTq4q9+T2iSmpfaE1zNfGs9D5W6d0aPtDQ7zcmt3lO2trSlKo3OSSlhZwj6z"
    "ox6wVaOZrojADvQeK9SOz6HDrRdSlb3kKvPW8KWHJezx9xqeLhnkZIpEf/j3cm/0azO2M3ax7Y+8"
    "OPOuXu1ds31az2tbVJU6sYyw54eMoxns6NnbUp2cJVKlTr4cqnUmhGXg0qTiqbcFzuP8z92XadKp"
    "N8tKKzH1PonH4+LBSIpVx2bkXyWmLSp8WpSo/ZvgVzHqunzfRHq/BnsucWePF7FaBYVdOiv++rRc"
    "Yte5z+x7wf0zjpxYo6FrdXw4afUVX6TSfkbrNt7a2/s7S7TR9F0mlawt4RoqdKkk3hYy36mn6n1X"
    "9t9lK+ZbDhdN/cfdafDWLr/dp702fsbU9z6/rdvVq6dQlXny1E8pGFl24Zq2dr0nQqunKfvhm4rt"
    "6cX7fhlwrudFu7/lnrtKdGmk8Sfp+RpqpSap1qc/mqVnU5vV5MnRs3I5de/JLH1Lj4uLPbRW4OeG"
    "3lrzZ2bhfpq1firt7SJR5o3FenFx9+p1unilDw5Pz9T1bsu7dr63x72td0KbqU7a5pufT6mz5uT6"
    "WCbPBxqd2aKt23DbQKW2Nlabo8IKEadCL5fvR2Vz5qeV6Fi7rWtlThWuakaVKEGnl4SR8qG+9nSk"
    "6cdatMx814i6HzW9u+83/Lv8VOykV/D7ra5MP1On8V92W+ydiavqlSfJKNtU5Jf7XKfUnvXarl4X"
    "7Zs05eX8VGP/AG/NwXulcEalzpkualXclzweU00Y59rtOnE/dl3vre2tarqdadVRupcnM8vGTrNz"
    "XnZ2nj27xTTWclVZzd3cV2syqycpfmcTVeerol030jFrIWbIO6t4JQstUu+J17bQq0rqlLwpVI5S"
    "l6YNg/FjZtxv/YmpbWt1TjVu4OMHLyXR4PB+7lsbe07Mmj3cI/67L5sdeiMid6b521w+2zcbv3Vf"
    "qz0u1ipVKsl5Z8gifbVnqvdLcbb/AHFf6nS3fYwtrmcpU4ePFcqb+84dp3SPHGjq1lc/vhZKnbVF"
    "OX8ZdVn7zPO37efZwuYqpS3inCXyvlXX/E+hoXbV4A7j1anoum7s57mrLlhFR83+YPLuXA7h9qPD"
    "DYdntXVLpV7mhCKc1LKbSwdl3VuvSNm6Bd7k3JeU7SzsKM6k4zml4mF6Z8y5r26ND29oFTdWt3sa"
    "GnUaSrc8+nRrKNQPbc7ZOu8a9xVdmbWvZ2miabOcFOi2lWWfXHmERDg9svtpbl477mutqbTvK1vt"
    "G2qOMaanhya6Z/wMVlTpum3GryRj5ym/NoZhKVR2qScFz1fr7nsnZf7Mu7O0ru63tbWyuKG16FZO"
    "5ruLXVefULuX2Uuyvu7tTb1tZfYqthtXTKkf2hKomo14p9cM3acO+G21+Fuz7LZOx9Pp2emWcOWK"
    "j0bfq2cXhfwv2vwj2fY7S2xp9K3jZ0YwnVpwSlVa9W/U7mqsYwU6nTPl0CkyupYSRJC6rJIAAAAA"
    "AIftgkiUlFZYRDzDtF6L+3uCG7bCcOecrCq4dM9T8+e59Jr7co6hpdxmM5XE8p9OmT9I+49Khrug"
    "Xuk1EnG6oyg0/qaC+2PtT92ONesaBSp+HTouc8YDJVsJ7pnZ2kWnCu51+lSh9plUxnHXqZ8qKcud"
    "vq0a7e6T3vZ3GyLrbsKy51NrkybFWlzP2SGvPdJb7Z0h9fh9ilRX83kTNSmk6Tx7kTpzlJJv4fUm"
    "J36YZru205Snz5xBJmtDvQ+A+6bqdLiXtK0q3niTzVp005YRsrqQnUxTj0h5M4Ou7d0jcWnT0fV7"
    "OncWs1iUZrPQz9P5U8PP3yx8vD9Wmn54bLR946za07a229XlUhHllFUn8x7LwD7HPFXi1u2wobk2"
    "5c2OkOcXKrOm0uXJt80zs2cH9KuJXNltqhGUpcz6Lz/I9EsNNs9Kt6djpNlQo0Ka5UoRSwbzlfqG"
    "166o1eHpMVt3S6zwp4dabwo2TYbE0mUZ0LOGE0sex3Vrojiwq20q7o0KydaPzLPU5Sfo/M5zJacv"
    "3T7luKUjHGoEunkUyXmy4US88e5Wn4Wmvd7eT9p7cctr8Fta1uM+SVKljOceaZoj1C+/ae6tS1yp"
    "Lmd3VnLPv1NxfeK7m/Y/Z91OwpVOWdddfwNM2jU3c2MKrfXLO1/TlNYZmfy5rrFv4moc6qlGlGpH"
    "1KKtNNUKk+sYVFLH3FVOUaj8GUklH1Yc6DbU7inyxT9TpckRNfLQYpmJluA7vPiHV3jw3/Zc6vPD"
    "TacYxWc49DLGEn4km/LLNY/dV8R6Ol3msbZvk8XclGnL8TZu4PkjT5sNtvJ836rXt5Vna9NneCNq"
    "03zFQ5UkiMtmur6e+lZjzKmfz0/73/I1W96vOvV4q6NbVFKNs7eGZenkbVJ9MS/p6mIvb67Pa4r7"
    "Cr7o0q3dbU7GmuRRWZNI93S89K8iJt6eHnUmcfhqG8OXiShTn/Dg+j9yE6k6nI+kfc5F9o+uaJN6"
    "Rqem1bSrbSanKqms4LDrU6tLwVWpqS83k+lfWwzjiYcXnpbu9O0cL+IOrcIN92m9tGm6s6E0217G"
    "wD/SqaZZ7Zt60Nrq6voU0qiTeW8GteNVU6btqDhJS8/U+ptzZ+7tyX9PR9t6Jc1q101GNSMG4xya"
    "zm8Pjcu8Xv8AD1cflcjD9lNw9K4/9p/dHaJ1upeapd1bOwUswsXJ4R5ElHk8NeiwZcbz7HlLgjwV"
    "ocRuIChWv7yCxBecG1nqjEiknUu6s4QXgzfND6Ivxrce2qYpjx4Ry+NmtH1LIoJ+Il7Myp7sPQKW"
    "4OO99O5pqcbBOssrPl1MWKLfjPlgnhmYndTVI0eKe8b6WIu3sak1n7mV6zMYePP058r9P++dWj0+"
    "N3oO5o65x00+2tpc1OwgqLw+mV0MU61FKTqp+uT0jtFbyrb74t7gua9GUnaX1SCk/pI835a9Wr4S"
    "ovD9TL06lONh7Ilj5Hdm3eYWrjKs6tVdZODwbRO7doWW0+z1q+/dVpRc6ClNOSx5Js1g0I1amqW+"
    "lOg348lBfXJs/srGtwe7AGqXzX2ercUuZJ9Hho1/WssWimOJ8TMRP9ns6ZjtXHa+vMMDe0bxf3Xx"
    "k4nat+2L+rW061rTVpSc3iKT9Dy9ywlHGZLzQt7+51ei9UjFOrcuT5vxHO1T8BU0rpfNI2nFpix0"
    "iIa+0ZMtp7oVeNHHPSo8iXRx9zufCriPa8Ldcp7nWmK71ChJTt6jWXSaOmKo6Ucyppy8mQpyin4d"
    "NZn55M1rYrR2ywxjyVnu0yq1PvGOJV9UjcwnVlWS5W02uhxod4bxPVVTVavL3hzMxgg50VmFOOX5"
    "9CKdSrSrqt4cM/d0Ndfp3HjzMQ9lefnjxG2V+h94BxJ1PcdpY06teVxWkowpcz6tnXe1jU3HuK0t"
    "95bjVWlfXKTcJ5ykzyDg1bw1DjntijGjGVSpcU8LHTzNkXb27O11vDhFY6jtOxzqllSjKtGnH5ko"
    "pvojxd3H4eWsfl7onNyqf2aqorkVOMZ88pRyVqnJyz4/I15orlZ3unScNQ0+paVLP+HU8RNZaLU6"
    "1k14sqsG5ezN93441FZaXLW3dt6JwO4z3/AbeNHdGk2ruKkprxMeqyZ03neuaLbaLCtDafNeSp4c"
    "cv5seZrVouVNeLbzhUT/AJfNn3Nr7A3bv/Vrey0bQbqUqtWMHUVN8uGzWc7h8XJaL5fL28blcisd"
    "mPw7Lxp46br437nrazuS/rV7GpNyt7aUnigm89Dz3njGo404c6XqZSdozs26b2duHWh6vq9KNW+1"
    "anFyS84toxZ55yjzWtPHM/I9XHy4fpR+3mNMWfHlmd5lyMPtGZ+XKn/kZh92bs+G596Xuqzoqo9O"
    "mpJ4zgw7jKca3hKKWYvP5GwPuiPCjc7rVRJTl8p5esZaxxLaln4GKZ5Mbj8Mue2LvO52PwlutXtr"
    "p0KlSEqcZKWMPBpVueM/GKhql/cWu87qdOrVcoJVX0WTZ33ru6rjSuCllZWc8Tr3DbWeuH0NTFrO"
    "rStqEY2cnOrTUm8epwE+odzSPtl3Ww4mcb9f1a1tqO+7unUqTilHxn7m1HdO1Nybu7EltZ6zUnqO"
    "qULbxJzl1lhR6s1Q8INP1PU+Mu3dPVpKrRq14KaSfTqb9tpbYtbXYNttu4op291aKlKDXTEo4Zjj"
    "2pWX546zq0tS1LT5U5PwKrhKTWOXqca9oKNjUtXUzGs1kzC7cfZA3Bwp3Tcbr2ZplS50i+m61aFK"
    "LeG3n0MQm4W/TV6X2fmfyVOjRK7Y72C+2ZtHY2waPDfdtSna21gm6dacsJ+55326u21DivKrwz2n"
    "NrRqvw1Zwl0nj1MI632WquWlf+HRj1XJLDLlvbxqyVxRhK4UenL5ybCFu3s7bT7TlqzlCjQfz87M"
    "uu7t7O9/xN33c74121qW+39E/wCswu55UZ8vXGfwPKOAHZX4g8f9zUbf9kXNhovPHxnUi0pRybQe"
    "KOm6P2TOy9Lau16dO3lUoSoVKkejlLlw+v4gYb9uztk6hvzVq3Dbh7fSpaVpUvslw6U+k+Xp1wYT"
    "RTUedNutluTbzzMqubh09X1XUvE8SN9XlVnzPLy2Q4SildxcceaQWiHZ+F+2NI3zu/T7HXNRjpVr"
    "RrRdxKTwpwz1RuH4acYOyxwV2jY6RtXcmm2nLRjGuoSinOeOrf4mk+vKVyvFozdKo11cG0yxHTY3"
    "D5bnUrt//lYTpvjl2zeA0baV3U3laRa8l4q6nfeHPF7ZnFO1+07Zv6d1Bdfhlk/OpqOn0o6Zczq6"
    "hdZpNcn8R+5uf7tDZFppPAfT90ylUnXvE0nN56LzDHMMxVjyXoSQvV+5IQAAAAAIZQpSllY6Fwp5"
    "kpcuAI8qkV/ss09d5jw5encX73dkqPJRvcx5sdOpuDnnnU1/Kma/u9t2q63DKx16yo/x3W5ZSS69"
    "AME+xNx4j2feNFhQ1u4dPQ7ypFTk30yzexom4LHc2jWe49BuqdxY31OFSnKDzlP7j83d1pVC/wBP"
    "tqFxU5biNNSjUzhqRkj2Ve8A4k9na7pbV3jGvrGiSapUnUk2qcPoTPidQm9tR5bx/Eisxgnze2Ca"
    "bnOP8RYZi/w/7ffCPedrRrXOqW9pOcU5RdRLDPVLHtC8INUpqcd6WUM9fiqpHo/a5pr3RV455eOs"
    "+ZemSdSLxGPT3KZzmscsW/wPMdQ7SXBrSYS8Te1lNx9FVTPJt994Pwn2kqqtr6hduC+HE/MnH07k"
    "Zp8VW/eYfyynlJeHmVSMEurb6YMYe152zdldn3aV3a6fqNK43BUjy0KUJZ+L8DC7jX3mm7N4yr6N"
    "sa0qWsJpxjWpvBhxuHUdy71117k37qVbUaiqKcYVJt+uTb4ehZK+bvBk6lG9eobtOxru/cXE/hXa"
    "8Ud0eJC81KUnGm0+kf8A9ZkJ4meqT6/QwR7MvbN4UbK4MaXt3VdQt7S5socvguaTX4Hpv+kI4IxW"
    "Hrltlf2iNdyuHmxWmZr4ezjcmmSI8soVNPzT/IicsyjhPz9jFuXeGcEE/wDtu2/4iJXeG8EOn/xu"
    "2/4iNdWt4nzD2WiLRp5v3nOoV6Gw5W1WM/s0oLmeOhqt0ura09KhOhPMW36mw/tydrLhhxY4Q3Og"
    "7bnQvL6axBxkm0a1tBoSt9OpUripicm/gZvOL1W3EpFY+GszdL+rM2l7/wACuzxuTjtcystqQ8Wp"
    "TTlUx5peplhsfuwadxXh+91epSin8Z5f3Uu8KujcX9V0nULtU7e4pShBSfRt+Rt3WZtuU4uMn8PT"
    "0wTn/Uua/iNaYqdJpXz8vBuD/Y64ccIKtG80CdT7RTw3LHm0e8um5csW2lHpkKCz0SWCpTSfSWTS"
    "ZM1+RPdLaYsEY41C5/LhEJe4XRZDfkV+GRLimmn6nFqW9tGnKjWpqrTn0cJJNP8ABnJcvoU4j6rm"
    "FftRMbeCcWexrwp4r1al7e2MbKvVy26UEllmPmr913tJXEnp15NwzlPJn7yRfnFFMoyfSE0vwPdj"
    "52XHGot4eS/HpafTCXZXdncOrOoq+4bio+V5wurZkZw27PHDjhVbShoGiUbip05alWnFyX3Hp/Km"
    "sVMP8BGCg8pJIpfm5Le7LRxqR8OgcWeD23eNO2Y7Y3U5wtYy5lTil0PFaPdzcEKNLwYyuOXGPlX6"
    "mU7prn8XPMn6B0qf+rwssx15V6eKTpacMT/NDFih3c/A62b5ZXHxeeYr9TuPCPsf8NeCmsahrG16"
    "1WM9TpulVUkl0Z7rGhRh0mk2KkI0oLFJSy/Qv+9y292mURx6R6qxq1TsBcGNU1K81WtGr419Vdao"
    "+VebZZj3e/BOMOVRrZ9+VfqZPRo02s8i/IOnTTxyIfvc0f75P29P6WMVv3fHBS3vra/jCs6lrNTh"
    "8K6tfietcRODG0+I/Dd8KdX5qOl+HGniC9EeieFD0gi3VpQq/wANrE/RkTysl9TayYw1rGqwxT0r"
    "u3uB+l2NKwhO4cKPyvlWf8y5Pu5eBrq+LKdwn/dX6mU9GEJZpuCzDpknkhKfJKisL1wW/e5v65V/"
    "b0j/AGsWf9HNwPf81w//AEr9R/o5eB/9Vx/ur9TKnwaX9CHgU/6UI5ub4vKJ41P6WKz7ubgh6yuP"
    "91fqQ+7m4HTjyRnccvr8K/Uyp8Cn/SvyKPDp58OnSWH5stPNzz7vKI42P+ljVtbsBcG9n7q0/d+m"
    "Ov8Aa9Omp0+aKxlfiZKXVC0uraVK5t41qM48jjJJpor8Ki3hJfD5iM1OTp+H8K8jy5OTa8/dO2eu"
    "KIjUQx14ndh/hJxL8evVtVZ1az5n4UUlk8H1HuvNs0a7pWF1OdFvo37GwKNOM84golUZNfAmun0P"
    "Zi6lmr6nf93lngY7T5hhdsru0OFukqFbWburUmv5V1MjuHnAnh/wztI22h6HbzccYqTpx5kehypp"
    "rMoKTf0KfDxDw1LDZiz9Qy3/AJrMlOFjr6h5fxq7O+y+O9C0t94Oo6Vn/qoRSaR5RDu5uB8KrqQl"
    "cZ9uVfqZTScoKMHT5/dk+DST55RSTMeHm5IjVbaZJ4tJ82jbFaXdycDnV8eU7hPH9K/U9L4KdmPY"
    "HAe5uLnZlSpzXXzqWOv5Hrk6VNzWIJxKpUYRg1CCi35GS2fJeNWtMqVw0ifEPL+NnZ12Xx5saOn7"
    "0lU8ChLmhGCT6nmVPu8eBtJ08UarjThyLMF5fmZOuj/BUKsstdcsmMYTXJKmkeW0vVXxDHrYnYZ4"
    "ObA3DDcul29Spc05KUHOC+FmQ1GEacI04RXhwiox+iRMaVKHToTNNpRj0QiCXzde2/o26LCppeta"
    "dSvKFSLi1UingxN4s92vwk3/AHNbVLSc7S4m3JU4xSjn6GYbqKOI0483uJxxicYrP3EjW/bd1Lo0"
    "Lnlldz8FPo8+h7DsDu3OE+2JUbzUJTr16bUuTlTj+JmDmbSeUvwLNS3dSvGt4mIr0A+TtrZ229o6"
    "bS0vb+j29pChBRUqVNJvHrlHV+NXBjQuN22Y7Z3DWlToxlzdD0Tmik+RZKOtX0xgDDmHdhcGF81z"
    "Xfv8Jej3ZPBdLDuK+PblMxFTjjrFEckc/IvyAw7fdk8Fs9Liuv8A0opl3ZHBiVZSVzX8P1XL1Mxn"
    "Thh/CvyLShHrSik0/UDDy57sLgfeRdGdxX8PKylFZMmeGXDjQ+E2y7LYm2E/sVkmocyw+v8A/DtX"
    "hJqVNU0vrgRpRUVThPE15gX15LPsSQuiSJAAAAAABQ8eJ5FZQ38aQENZ5l7ox/7bWwae/OD15bzo"
    "+K7RSq4xnHQyBn5o6/v7To6nszXbOUFPxrOpFJr15QQ/Old0KcNT1C1qPLtKrppL0wy3cc2oKFvc"
    "xi6S6dI9Ts/Evbv7n8Qdfs7mMo+PdzaT/vM61WUrReJQ+LmJpaKZI7ls9O6vhfpaH9njGWk3txTf"
    "spNHOhW3ha4VPW7pR/8Aqs4tC6nSt1WlLEn6FEr6s4fxJvD+p2uHqXGxcfzENJk4Fry5lWWu3Kbu"
    "NZupS9f4rOE9OtKs86pd1ppP1lktxq3Di5Upt/iUUIzuJtXUsYNVHX6RaYrC+Lo9vfc+g7rTral4"
    "ej01GS9ZItXGpSnQ/jrNVLpg4lSNKnJql6EqMnHmaX4ngt1bPbfl756fgfNutIncR+2VLyvGUn1j"
    "GWEcb92bSfV3lzl+fxs+3GPO8TnHHtzFFWtQoR60qs/7iyeDLycuafMs+PBix+nyP3Rspdfttz/v"
    "sfulaZUVd3XXp8zPtaZTvNYuI22maZeyqSeF/DeD1Hb3ZV4/70pQr7c0CrKnPquaBWlLTG5llmKv"
    "HrHSf2FcKtK8qVKT/lnLKObKrYVK/wBo+00m38sUzLnh/wB3Bxs1qrCnu7Tp0aUvmePIyL2V3TWx"
    "aMqV5uPUqsaiw3BMpSdz9ysZZlgN2feIGqcNeKGianb6XXrw1C6hScqUW8JyRv00e4ne6HpV4k4u"
    "4tqNRp+a5oJnh2yexHwb2Z9lnR06NzVtJKVOVSCeGvUyAoUKdvRpW9NJU6MVGK9klhF7Up8QTO1b"
    "wopSKYwgusUVvr5EYeSsWiPEMU9ypYawQ1gRKidphQIpYaj5kvzIE+UrcY1U3zvKKlFecSvLGfoj"
    "HNJNKeWMnnHVEOXPFwRcePIjES/gW4x5YeH6rqThTlzpYaLiSHKhqDaz4acuap1Ko86ln+Urwvcn"
    "KQ8J2ofN6E4z5k4XuML3J8G0cz6pFEIvPPU6y9C7he5GF7jwLSbnl0lh+pU+dxSb6leI+nQlYXqP"
    "CELGOq6kepVlEPr5FfAZX4FuWUuSn0b9S5y+5OEW3AtU4xp9MfE/MrwvlSwVYiMrJExEm1OeXqyF"
    "GMk3HoyttepGceQrWK+keVEPEgv4jyJQU/ij0ZXljC9xNYt7NzChczaS9PMmUFUliXkirC9wsL1E"
    "Vivo3Knl5ekfIKMpJ5/AuATJpainJctTq0Vywvia8ioEJW1Dq5yJk5SjmJWC0eESt0uXlfIsfeVZ"
    "k08E4SJCVtRk3ifUP51D0LgAtqLjPEPL1DlmTjFdV6lwjC9gHxe4+L3JAEdfX1LU3yfw6SxJ+peI"
    "wvPAFun4sViq85JjCmqjml8TLgAAAAAAAAAFEliWSsoXxLqgKms9SxcUVcWte3mulWMoPP1WC+vY"
    "o6uovbIGkHvAtpWm0+NzsoUlTVzJz8sepjpGcYutSl1UW1E2Fd6lw0tKu4bXe/NCj4VNOU5dFk15"
    "2Vvqt9QdbTtHr3qk3iVKLakUyY5v6Ze+LLcKbqRaqVOWPoVq2dODWXP2OybR4YcQ963sbCjsvUaE"
    "ZPCqSpSS/wAj37aXdo8cN8RhXsr9WFOXXNV8qS/EiK212ypNJ+GLMbhW2XXcqcV9CmleWV5U5KFy"
    "5TfRJR82bI+GvdT6zp1SC3/q9te001zKFRMyT2t3evZy29SpSvNtRuLinhuTkl1MmPFjp5lG7x8t"
    "NelcN+J2uNfu/tK4vef5eWk3k9J2R2POPW8bqFDU9mXdhRqNfG6TWEzdptTg/wAOdlKC2/t62pOm"
    "sQzFSx/gdwhzwSjChTjFeSTwROpVike5ardqd0hrGtW1C81jdErNvDlBmQnD7uzeHO0qlGpq95S1"
    "PkxzKUPP8zM+XNUWEsfcyYUuXzlJ/iTFYhbUQ8s0Hs0cD9twhGz2Dp7qQS+OVHLbPQdK2/oOhUlS"
    "0bTKFpFdEoU8H0pRm38LJaUV8SyJma+iJgU5tZ6fkMuXRpP8AkkV4WPIiLxKZ1C06UW88qK8Mny8"
    "0Rl+5M+YV7jqhl+4C+pEQjuVY9SQCUowiOX6jLGWV7g5fqQ+hUnkYRMTsQ+rHL9RljLKxUThr1GG"
    "/Upy/cZfuW7U6Ty/UhrDJyxllZrKDl+o5fqMsZY7ZDl+o5fqMsZY7ZDl+oaGWE/cdsiCUsk4T6jK"
    "Q7ERKM46Dm+hGX7k5Y7JSgLqycsnPTI7ZENdSOpLfXITLan4R3wjD9gTn1Jwn5iIn5TuFIKsIdF6"
    "ETtO4SCCSYhXYABo2AFLbySlUAAAAAAAAAAAAAAAAAAAAAAAAAABGFnJIA8l4/8AZv2R2iNA/d3e"
    "NavSo/1UUuY61wv7F3Bfhfp9tpmlaZC9jbdE68Itv7/M95r0eem44yy34dGhBQUVGUljIHzbHaW1"
    "tOiray2tp9GEPJxtof54yfXp0KVKHJQpqivaMUkRFRcfDjUzJF1dEBSoTXnVb/BE8ifn1KgBGF7D"
    "lRIAjCRIAAAAAARoCML2JBIjC9hhexIAAAAACNQAAJ0AAAjCGESAAAAAAAAAAAAAAARjpgkAAABG"
    "F7DC9iQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAFurGq2nTeC3UlTUowqxbk/U5BD"
    "SfVpAW4U6UKj5U+Zl0jC8yQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAA"
    "AAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAP/9k="
)

from reportlab.platypus import Image as RLImage

# Barevná paleta Alza
ALZA_BLUE    = colors.HexColor('#0A3D91')   # tmavě modrá (logo)
ALZA_BLUE_DK = colors.HexColor('#072B5E')   # tmavší modrá (akcenty)
TEXT_DARK    = colors.HexColor('#1A1A1A')
TEXT_MUTED   = colors.HexColor('#666666')
LINE_GRAY    = colors.HexColor('#CCCCCC')
ROW_ALT      = colors.HexColor('#F8F9FB')


def _alza_logo_image(width=2.5*cm):
    """Vrátí obrázek loga Alza.cz z embeded base64 (čtvercové logo 600x600)."""
    import base64 as _b64
    logo_bytes = _b64.b64decode(_ALZA_LOGO_B64)
    logo_io = io.BytesIO(logo_bytes)
    # Poměr loga: 600x600 (čtverec) → výška = width
    return RLImage(logo_io, width=width, height=width)


def _hlavicka_alza(titulek_text, W):
    """Hlavička: logo vlevo, nadpis vpravo, pod tím horizontální linka."""
    title_s = ParagraphStyle('t', fontSize=18, fontName=PDF_FONT_BOLD,
                             textColor=TEXT_DARK, alignment=2, leading=22)

    logo_img = _alza_logo_image(width=2.5*cm)
    ht = Table([[logo_img, Paragraph(titulek_text, title_s)]],
               colWidths=[3.0*cm, W - 3.0*cm])
    ht.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    # Horizontální linka pod hlavičkou
    linka = Table([['']], colWidths=[W], rowHeights=[1])
    linka.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, -1), 1.5, ALZA_BLUE),
    ]))

    return [ht, Spacer(1, 0.3*cm), linka]


def _alza_spolecnost(W):
    """Informace o společnosti Alza.cz a.s."""
    s = ParagraphStyle('c', fontSize=9.5, fontName=PDF_FONT, textColor=TEXT_DARK, leading=13)
    b = ParagraphStyle('cb', fontSize=9.5, fontName=PDF_FONT_BOLD, textColor=TEXT_DARK, leading=13)

    data = [
        [Paragraph('<b>Společnost:</b>', b), Paragraph('Alza.cz a.s.', s)],
        [Paragraph('<b>Sídlo:</b>', b), Paragraph('Jankovcova 1522/53, 170 00 Praha 7', s)],
        [Paragraph('<b>IČO:</b>', b), Paragraph('27082440', s)],
    ]
    t = Table(data, colWidths=[2.5*cm, W - 2.5*cm])
    t.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 1),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 1),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    return t


def _paticka_alza():
    """Diskrétní patička."""
    footer_s = ParagraphStyle('f', fontSize=7.5, fontName=PDF_FONT,
                              textColor=TEXT_MUTED, alignment=1)
    return Paragraph(
        f"Alza.cz a.s. &nbsp;·&nbsp; Facility &nbsp;·&nbsp; "
        f"Dokument vygenerován {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        footer_s)


def _pravni_text(W):
    """Právní prohlášení — diskrétní, bez barev."""
    legal_s = ParagraphStyle('leg', fontSize=8, fontName=PDF_FONT,
                             textColor=TEXT_DARK, leading=11, alignment=4)
    nadpis_s = ParagraphStyle('ln', fontSize=9, fontName=PDF_FONT_BOLD,
                              textColor=ALZA_BLUE, leading=12)

    elementy = [
        Paragraph("Prohlášení zaměstnance — NV č. 390/2021 Sb.", nadpis_s),
        Spacer(1, 0.15*cm),
    ]
    legal_txt = (
        "Předání a převzetí výše uvedených OOPP a předmětů zaměstnanec i zaměstnavatel "
        "potvrzují svým podpisem. Zaměstnanec byl seznámen se způsobem údržby OOPP "
        "dle nařízení vlády č. 390/2021 Sb. Zaměstnanec se zavazuje řádně hospodařit s OOPP "
        "a předměty svěřenými mu zaměstnavatelem na základě tohoto potvrzení a střežit "
        "a ochraňovat tyto OOPP a předměty zaměstnavatele před poškozením, ztrátou, zničením "
        "a zneužitím. Zaměstnanec se zavazuje svěřené OOPP používat výhradně pro výkon práce "
        "pro zaměstnavatele. Zaměstnanec souhlasí s tím, že v případě ztráty nebo poškození "
        "bude cena sražena ze mzdy v souladu s příslušnou Dohodou o srážkách ze mzdy."
    )
    elementy.append(Paragraph(legal_txt, legal_s))
    return elementy


def _podpisy_alza(W):
    """Podpisová sekce — dvě linky s popiskami 'Předávající' a 'Přebírající'."""
    label_s = ParagraphStyle('pl', fontSize=10, fontName=PDF_FONT_BOLD,
                             textColor=TEXT_DARK, alignment=1)
    date_s = ParagraphStyle('pd', fontSize=9, fontName=PDF_FONT,
                            textColor=TEXT_MUTED, alignment=1)

    data = [
        # prostor pro podpisy
        ['', '', '', ''],
        # linky
        ['', '', '', ''],
        # popisky pod linkami
        [Paragraph('Předávající', label_s), '',
         Paragraph('Přebírající', label_s), ''],
        # datum
        [Paragraph(f'V Chrášťanech dne {datetime.now().strftime("%d.%m.%Y")}', date_s), '',
         Paragraph(f'V Chrášťanech dne {datetime.now().strftime("%d.%m.%Y")}', date_s), ''],
    ]
    t = Table(data,
              colWidths=[W/2 - 0.5*cm, 1.0*cm, W/2 - 0.5*cm, 0],
              rowHeights=[1.5*cm, 0.1*cm, 0.5*cm, 0.4*cm])
    t.setStyle(TableStyle([
        ('LINEABOVE', (0, 1), (0, 1), 1.0, TEXT_DARK),
        ('LINEABOVE', (2, 1), (2, 1), 1.0, TEXT_DARK),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
    ]))
    return t


# ═══════════════════════════════════════════════════════════════════
# PDF PROTOKOL — MCDP (mycí a čisticí prostředky)
# ═══════════════════════════════════════════════════════════════════
def generovat_pdf_protokol(zamestnanec, sklad, kvartal, vydane_polozky, vedouci, velikosti=None):
    """Předávací protokol MČDP ( Mycí a čisticí prostředky ) — čistý Alza styl, originální logo."""
    if velikosti is None:
        velikosti = {}

    body_s  = ParagraphStyle('b', fontSize=10, fontName=PDF_FONT,
                             textColor=TEXT_DARK, leading=14)
    body_b  = ParagraphStyle('bb', fontSize=10, fontName=PDF_FONT_BOLD,
                             textColor=TEXT_DARK, leading=14)
    section_s = ParagraphStyle('sec', fontSize=11, fontName=PDF_FONT_BOLD,
                               textColor=ALZA_BLUE, leading=14, spaceAfter=4)
    th_s = ParagraphStyle('th', fontSize=9.5, fontName=PDF_FONT_BOLD,
                          textColor=colors.white, alignment=1)
    td_s = ParagraphStyle('td', fontSize=10, fontName=PDF_FONT,
                          textColor=TEXT_DARK, leading=13)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=2.0*cm, leftMargin=2.0*cm,
        topMargin=1.0*cm, bottomMargin=1.0*cm,
        title="Předávací protokol MČDP")
    el = []
    W = 17.0 * cm

    # Hlavička s logem
    el.extend(_hlavicka_alza("Předávací protokol — MČDP", W))
    el.append(Spacer(1, 0.4*cm))

    # Společnost
    el.append(_alza_spolecnost(W))
    el.append(Spacer(1, 0.2*cm))
    el.append(Paragraph('Dále jen „předávající"', body_s))
    el.append(Spacer(1, 0.3*cm))
    el.append(Paragraph("<b>a</b>", body_s))
    el.append(Spacer(1, 0.3*cm))

    # Info o zaměstnanci - inline styl (jako originál)
    udaje_data = [
        [Paragraph('<b>Jméno a příjmení:</b>', body_b),
         Paragraph(zamestnanec or '…………………………………………', body_s)],
        [Paragraph('<b>Sklad:</b>', body_b),
         Paragraph(sklad, body_s)],
        [Paragraph('<b>Kvartál / rok:</b>', body_b),
         Paragraph(kvartal, body_s)],
        [Paragraph('<b>Datum výdeje:</b>', body_b),
         Paragraph(datetime.now().strftime('%d.%m.%Y'), body_s)],
        [Paragraph('<b>Vedoucí / zadal:</b>', body_b),
         Paragraph(vedouci or '—', body_s)],
        [Paragraph('<b>Číslo protokolu:</b>', body_b),
         Paragraph(f"MCDP-{sklad}-{datetime.now().strftime('%Y%m%d%H%M')}", body_s)],
    ]
    ut = Table(udaje_data, colWidths=[4.0*cm, W - 4.0*cm])
    ut.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    el.append(ut)
    el.append(Spacer(1, 0.5*cm))

    # Úvodní text — OPRAVENO
    el.append(Paragraph("Předávající předává a přebírající přijímá:", body_s))
    el.append(Spacer(1, 0.3*cm))

    # Tabulka položek
    el.append(Paragraph("Vydávané položky", section_s))

    def mark(val):
        return '✓' if val else '—'

    # OPRAVENO: ručník má False (nemá prázdnou linku na velikost)
    polozky_def = [
        ('1× Ručník Siguro 50×100 cm', 'rucnik',  '50×100 cm, froté',        False),
        ('1× Tekuté mýdlo',             'mydlo',   '500 ml',                   False),
        ('1× Ariel tablety',            'ariel',   '60 ks / balení',           False),
        ('1× Krém Indulona',            'krem',    'originál nebo měsíčkový', False),
        ('1× Abrazivní pasta Solvina',  'solvina', '450 g',                    False),
    ]

    header_row = [
        Paragraph('Položka', th_s),
        Paragraph('Vydáno', th_s),
        Paragraph('Velikost', th_s),
        Paragraph('Specifikace', th_s),
        Paragraph('Podpis', th_s),
    ]
    polozky_data = [header_row]
    for nazev, klic, spec, je_odev in polozky_def:
        vel_val = velikosti.get(klic, '') if je_odev else ''
        velikost_cell = vel_val if vel_val else ('__________' if je_odev else '—')
        polozky_data.append([
            Paragraph(nazev, td_s),
            mark(vydane_polozky.get(klic)),
            velikost_cell,
            Paragraph(spec, td_s),
            '',
        ])

    col_w = [5.8*cm, 1.6*cm, 2.2*cm, 4.0*cm, 3.4*cm]
    pt = Table(polozky_data, colWidths=col_w,
               rowHeights=[0.8*cm] + [0.85*cm]*len(polozky_def))
    pt.setStyle(TableStyle([
        # Hlavička tabulky - tmavě modrá Alza
        ('BACKGROUND', (0, 0), (-1, 0), ALZA_BLUE),
        # Zebrování
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        # Rámeček
        ('BOX', (0, 0), (-1, -1), 0.8, TEXT_DARK),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, LINE_GRAY),
        # Text ve sloupci Vydáno - větší, modrý
        ('FONTNAME', (1, 1), (1, -1), PDF_FONT_BOLD),
        ('FONTSIZE', (1, 1), (1, -1), 14),
        ('TEXTCOLOR', (1, 1), (1, -1), ALZA_BLUE),
        ('ALIGN', (1, 0), (1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        # Linka ve sloupci podpis
        ('LINEBELOW', (4, 1), (4, -1), 0.6, TEXT_MUTED),
    ]))
    el.append(pt)
    el.append(Spacer(1, 0.5*cm))

    # OPRAVENO: Přebírající stvrzuje (ne Předávající)
    el.append(Paragraph(
        "Přebírající stvrzuje, že se řádně seznámil se stavem předmětu "
        "a převzal jej v pořádku.",
        body_s))
    el.append(Spacer(1, 0.2*cm))
    el.append(Paragraph(
        "Pokud je předmět kód, čip nebo klíče, přebírající potvrzuje, "
        "že v případě ztráty je povinen uhradit částku ve výši odpovídající "
        "skutečným nákladům.",
        body_s))
    el.append(Spacer(1, 0.4*cm))

    # Podpisy
    el.append(_podpisy_alza(W))
    el.append(Spacer(1, 0.3*cm))

    # Právní text
    el.extend(_pravni_text(W))

    # Patička
    el.append(Spacer(1, 0.4*cm))
    el.append(_paticka_alza())

    doc.build(el)
    buf.seek(0)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════════
# PDF PROTOKOL — OOPP (osobní ochranné pracovní pomůcky)
# ═══════════════════════════════════════════════════════════════════
def generovat_pdf_oopp(zamestnanec, email, sklad, vydane_pomucky, velikosti_oopp,
                        expirace_oopp, vedouci, stredisko="", osobni_cislo=""):
    """Předávací protokol OOPP — čistý Alza styl, originální logo."""
    if velikosti_oopp is None:
        velikosti_oopp = {}
    if expirace_oopp is None:
        expirace_oopp = {}

    body_s  = ParagraphStyle('b', fontSize=10, fontName=PDF_FONT,
                             textColor=TEXT_DARK, leading=14)
    body_b  = ParagraphStyle('bb', fontSize=10, fontName=PDF_FONT_BOLD,
                             textColor=TEXT_DARK, leading=14)
    section_s = ParagraphStyle('sec', fontSize=11, fontName=PDF_FONT_BOLD,
                               textColor=ALZA_BLUE, leading=14, spaceAfter=4)
    th_s = ParagraphStyle('th', fontSize=9.5, fontName=PDF_FONT_BOLD,
                          textColor=colors.white, alignment=1)
    td_s = ParagraphStyle('td', fontSize=9.5, fontName=PDF_FONT,
                          textColor=TEXT_DARK, leading=12)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
        rightMargin=2.0*cm, leftMargin=2.0*cm,
        topMargin=1.0*cm, bottomMargin=1.0*cm,
        title="Předávací protokol OOPP")
    el = []
    W = 17.0 * cm

    # Hlavička s logem
    el.extend(_hlavicka_alza("Předávací protokol — OOPP", W))
    el.append(Spacer(1, 0.3*cm))

    # Společnost
    el.append(_alza_spolecnost(W))
    el.append(Spacer(1, 0.25*cm))

    # Info o zaměstnanci - dvousloupcová tabulka pro úsporu místa
    udaje_data = [
        [Paragraph('<b>Jméno a příjmení:</b>', body_b),
         Paragraph(zamestnanec or '…………………………………………', body_s),
         Paragraph('<b>Sklad:</b>', body_b),
         Paragraph(sklad, body_s)],
        [Paragraph('<b>Email:</b>', body_b),
         Paragraph(email or '—', body_s),
         Paragraph('<b>Datum výdeje:</b>', body_b),
         Paragraph(datetime.now().strftime('%d.%m.%Y'), body_s)],
        [Paragraph('<b>Středisko:</b>', body_b),
         Paragraph(stredisko or '—', body_s),
         Paragraph('<b>Vedoucí / zadal:</b>', body_b),
         Paragraph(vedouci or '—', body_s)],
        [Paragraph('<b>Osobní číslo:</b>', body_b),
         Paragraph(osobni_cislo or '—', body_s),
         Paragraph('<b>Číslo protokolu:</b>', body_b),
         Paragraph(f"OOPP-{sklad}-{datetime.now().strftime('%Y%m%d%H%M')}", body_s)],
    ]
    _unused_fix_udaje = []
    ut = Table(udaje_data, colWidths=[3.5*cm, (W/2 - 3.5*cm), 3.5*cm, (W/2 - 3.5*cm)])
    ut.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
    ]))
    el.append(ut)
    el.append(Spacer(1, 0.3*cm))

    # Úvod — OPRAVENO
    el.append(Paragraph("Předávající předává a přebírající přijímá:", body_s))
    el.append(Spacer(1, 0.2*cm))

    # Tabulka pomůcek
    el.append(Paragraph("Vydávané pomůcky", section_s))

    pomucky_def = [
        ('Oděv pracovní (montérky)', 'odev',     'dle standardu firmy'),
        ('Rukavice bezpečnostní',    'rukavice', 'dle potřeby pozice'),
        ('Kabát proti chladu',       'kabat',    'zimní období'),
        ('Tričko',                   'tricko',   'letní období'),
        ('Mikina',                   'mikina',   'přechodné období'),
        ('Čepice / kšiltovka',       'cepice',   'dle potřeby'),
        ('Ochranné brýle',           'bryle',    'dle pracoviště'),
        ('Kraťasy',                  'kratasy',  'letní období'),
        ('Thermo prádlo',            'thermo',   'zimní období'),
        ('Bezpečnostní obuv',        'obuv',     'S1P / S3'),
    ]

    def mark(val):
        return '✓' if val else '—'

    header_row = [
        Paragraph('Pomůcka', th_s),
        Paragraph('Vydáno', th_s),
        Paragraph('Velikost', th_s),
        Paragraph('Expirace', th_s),
        Paragraph('Specifikace', th_s),
        Paragraph('Podpis', th_s),
    ]
    polozky_data = [header_row]
    for nazev, klic, spec in pomucky_def:
        vydano = vydane_pomucky.get(klic, False)
        vel = velikosti_oopp.get(klic, '') if vydano else ''
        exp = expirace_oopp.get(klic, '') if vydano else ''
        polozky_data.append([
            Paragraph(nazev, td_s),
            mark(vydano),
            vel if vel else ('__________' if vydano else '—'),
            exp if exp else '—',
            Paragraph(spec, td_s),
            '',
        ])

    col_w = [4.3*cm, 1.4*cm, 2.2*cm, 2.0*cm, 3.6*cm, 3.5*cm]
    pt = Table(polozky_data, colWidths=col_w,
               rowHeights=[0.75*cm] + [0.7*cm]*len(pomucky_def))
    pt.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), ALZA_BLUE),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ('BOX', (0, 0), (-1, -1), 0.8, TEXT_DARK),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, LINE_GRAY),
        ('FONTNAME', (1, 1), (1, -1), PDF_FONT_BOLD),
        ('FONTSIZE', (1, 1), (1, -1), 13),
        ('TEXTCOLOR', (1, 1), (1, -1), ALZA_BLUE),
        ('ALIGN', (1, 0), (3, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LEFTPADDING', (0, 0), (-1, -1), 6),
        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
        ('LINEBELOW', (5, 1), (5, -1), 0.6, TEXT_MUTED),
    ]))
    el.append(pt)
    el.append(Spacer(1, 0.5*cm))

    # OPRAVENO: Přebírající stvrzuje (ne Předávající)
    el.append(Paragraph(
        "Přebírající stvrzuje, že se řádně seznámil se stavem předmětu "
        "a převzal jej v pořádku.",
        body_s))
    el.append(Spacer(1, 0.2*cm))
    el.append(Paragraph(
        "Pokud je předmět kód, čip nebo klíče, přebírající potvrzuje, "
        "že v případě ztráty je povinen uhradit částku ve výši odpovídající "
        "skutečným nákladům.",
        body_s))
    el.append(Spacer(1, 0.4*cm))

    # Podpisy
    el.append(_podpisy_alza(W))
    el.append(Spacer(1, 0.3*cm))

    # Právní text
    el.extend(_pravni_text(W))

    # Patička
    el.append(Spacer(1, 0.4*cm))
    el.append(_paticka_alza())

    doc.build(el)
    buf.seek(0)
    return buf.getvalue()



# ═══════════════════════════════════════════════════════════════════
# KONFIGURACE STRÁNKY
# ═══════════════════════════════════════════════════════════════════
st.set_page_config(page_title="DocScan", layout="wide", page_icon="🔍")

st.markdown("""
<style>
    [data-testid="stMainViewContainer"] .block-container { max-width: 1200px !important; margin-left: auto !important; margin-right: auto !important; }
    .stApp { background: linear-gradient(135deg, #051c3d 0%, #2e0b54 40%, #1a0633 70%, #030821 100%) !important; background-attachment: fixed !important; color: #f0f0f0; }
    [data-testid="stHeader"] { background: rgba(0,0,0,0) !important; }
    div[data-testid="stMetric"] { background: rgba(0,0,0,0.25) !important; backdrop-filter: blur(15px); padding: 15px; border-radius: 8px; border: 1px solid rgba(255,255,255,0.6) !important; box-shadow: 0 0 15px rgba(0,242,255,0.4) !important; height: 90px !important; }
    [data-testid="stMetricValue"] { font-size: 1.5rem !important; }
    [data-testid="stMetricLabel"] p { font-size: 0.9rem !important; }
    div[data-testid="stButton"] > button { background: linear-gradient(135deg, #0052cc 0%, #0a84ff 100%) !important; border: none !important; color: #ffffff !important; box-shadow: 0 0 8px rgba(0,82,204,0.4) !important; transition: all 0.2s ease-in-out !important; font-weight: 600 !important; text-transform: uppercase; letter-spacing: 1.5px; height: 38px !important; font-size: 0.75rem !important; border-radius: 8px !important; }
    div[data-testid="stButton"] > button:hover { box-shadow: 0 0 14px rgba(0,82,204,0.6) !important; transform: scale(1.01); }
    .energy-card { background: rgba(10,10,20,0.4) !important; border-radius: 12px; padding: 6px 10px; border: 1px solid rgba(255,255,255,0.1); backdrop-filter: blur(20px); margin-bottom: 6px; }
    .energy-card h3 { font-size: 0.9rem !important; margin: 4px 0 !important; }
    .el-border { border-top: 2px solid #FFD700 !important; }
    .fsx-border { border-top: 2px solid #0084ff !important; }
    .gas-border { border-top: 2px solid #FF5722 !important; }
    .water-border { border-top: 2px solid #00BFFF !important; }
    .oopp-border { border-top: 2px solid #00c864 !important; }
    [data-testid="stFileUploadDropzone"], section[data-testid="stFileUploadDropzone"], section[data-testid="stFileUploadDropzone"] > div { background-color: rgba(0,200,100,0.08) !important; border: 2px dashed #00c864 !important; border-radius: 10px !important; }
    div[data-baseweb="select"] > div { background-color: rgba(0,200,100,0.06) !important; border: 1px solid rgba(0,200,100,0.3) !important; }
    span[data-baseweb="tag"] { background-color: rgba(0,200,100,0.15) !important; border: 1px solid rgba(0,200,100,0.4) !important; color: #00e87a !important; }
    [data-testid="stDataFrame"] { background-color: rgba(0,82,204,0.05) !important; padding: 10px; border-radius: 10px; border: 1px solid rgba(0,132,255,0.3) !important; }
    .cat-card { background: rgba(255,255,255,0.04); border: 1px solid rgba(255,255,255,0.1); border-radius: 16px; padding: 16px; text-align: center; position: relative; transition: all 0.3s; cursor: pointer; margin-bottom: 8px; }
    .cat-card:hover { background: rgba(255,255,255,0.08); border-color: rgba(255,255,255,0.3); transform: translateY(-2px); box-shadow: 0 8px 20px rgba(0,0,0,0.3); }
    .cat-card.active { background: rgba(0, 135, 90, 0.15); border-color: #00875a; box-shadow: 0 0 20px rgba(0, 135, 90, 0.3); }
    .cat-name { font-weight: bold; color: #fff; font-size: 0.9rem; margin-top: 6px; }
    .cat-desc { font-size: 0.7rem; color: rgba(255,255,255,0.4); margin-top: 4px; }
    .preview-row { display: flex; justify-content: space-between; padding: 6px 0; border-bottom: 1px solid rgba(255,255,255,0.06); }
    .preview-row:last-child { border-bottom: none; }
    .preview-label { font-size: 0.75rem; color: #888; text-transform: uppercase; }
    .preview-value { font-size: 0.85rem; color: #ccc; font-style: italic; }
    div[data-testid="stDownloadButton"] > button { background: rgba(0, 135, 90, 0.15) !important; border: 1px solid rgba(0, 135, 90, 0.4) !important; color: #00875a !important; font-size: 0.8rem !important; border-radius: 8px !important; height: 36px !important; box-shadow: none !important; width: auto !important; text-transform: none !important; letter-spacing: 0 !important; font-weight: normal !important; }
    .obdobi-box { background: rgba(0,200,100,0.06); border: 1px solid rgba(0,200,100,0.3); border-radius: 10px; padding: 12px 16px; color: #00e87a; font-size: 0.9rem; margin-bottom: 8px; }
    .obdobi-box-empty { background: rgba(0,200,100,0.06); border: 1px solid rgba(0,200,100,0.3); border-radius: 10px; padding: 12px 16px; color: #555; font-size: 0.85rem; font-style: italic; margin-bottom: 8px; }
</style>
""", unsafe_allow_html=True)

st.title("🔍 DocScan")
st.write("---")

if 'vysledky' not in st.session_state: st.session_state.vysledky = []
if 'kategorie' not in st.session_state: st.session_state.kategorie = "Energie"
if 'pocet_souboru' not in st.session_state: st.session_state.pocet_souboru = 0
if 'datum_analyzy' not in st.session_state: st.session_state.datum_analyzy = None
if 'obdobi_input' not in st.session_state: st.session_state.obdobi_input = ''
if 'file_uploader_key' not in st.session_state: st.session_state.file_uploader_key = 0

kategorie_list = [
    ("⚡", "Energie",      "Spotřeba & náklady"),
    ("📄", "Faktury",      "Dodavatel, částky, splatnost"),
    ("📋", "Smlouvy",      "Strany, podmínky, datum"),
    ("🦺", "OOPP & MČDP", "Evidence & výdej pomůcek"),
]

cols_kat = st.columns(4)
for col, (icon, name, desc) in zip(cols_kat, kategorie_list):
    with col:
        active = "active" if st.session_state.kategorie == name else ""
        st.markdown(
            f'<div class="cat-card {active}"><div style="font-size:1.8rem">{icon}</div>'
            f'<div class="cat-name">{name}</div><div class="cat-desc">{desc}</div></div>',
            unsafe_allow_html=True)
        if st.button(name, key=f"btn_{name}", use_container_width=True):
            st.session_state.kategorie = name
            st.rerun()

st.write("---")

pocet_souboru = st.session_state.get('pocet_souboru', 0)
obdobi_stat = st.session_state.vysledky[0].get('obdobi', '—') if st.session_state.vysledky else '—'
celkem_nakladu = 0
if st.session_state.vysledky:
    res = st.session_state.vysledky[0]
    for klic in ['el_cena_celkem_zaklad_kc', 'fsx_cena_bez_dph', 'plyn_cena_celkem_zaklad_kc', 'voda_cena_bez_dph']:
        try:
            hodnota = res.get(klic, 0)
            if hodnota and str(hodnota).lower() != 'n/a':
                celkem_nakladu += float(str(hodnota).replace(',', '.').replace(' ', ''))
        except Exception:
            pass

c1, c2, c3, c4 = st.columns(4)
with c1: st.metric("Nahráno souborů", str(pocet_souboru) if pocet_souboru > 0 else "—")
with c2: st.metric("Období", obdobi_stat)
with c3: st.metric("Ušetřeno času", "~10 min" if st.session_state.vysledky else "—")
with c4: st.metric("Celkem nákladů", f"{celkem_nakladu:,.0f} Kč".replace(",", " ") if celkem_nakladu > 0 else "—")

st.write("---")
col_side, col_main = st.columns([1, 3])

# ── ENERGIE ──────────────────────────────────────────────────────
if st.session_state.kategorie == "Energie":
    with col_side:
        st.markdown('<p style="color:#00c864;font-size:0.75rem;font-weight:bold;letter-spacing:2px;text-transform:uppercase;">Konfigurace</p>', unsafe_allow_html=True)
        sklad = st.selectbox("Sklad:", ["CZLC4", "LCÚ", "LCZ", "SKLC3"])
        st.markdown('<p style="color:#00c864;font-size:0.75rem;font-weight:bold;letter-spacing:2px;text-transform:uppercase;margin-bottom:4px;">Období</p>', unsafe_allow_html=True)
        obdobi_val = st.session_state.get('obdobi_input', '')
        if obdobi_val:
            st.markdown(f'<div class="obdobi-box">📅 {obdobi_val}</div>', unsafe_allow_html=True)
        else:
            st.markdown('<div class="obdobi-box-empty">Vyplní se automaticky po analýze</div>', unsafe_allow_html=True)
        uploaded_files = st.file_uploader("Vložte dokumenty", accept_multiple_files=True,
            type=['pdf', 'docx', 'xlsx', 'xls'],
            key=f"uploader_{st.session_state.file_uploader_key}")
        if uploaded_files:
            st.markdown(f'<p style="color:#00c864;font-size:0.8rem;">✓ {len(uploaded_files)} soubor(ů) připraveno</p>', unsafe_allow_html=True)
        st.write("")
        _, mid_btn, _ = st.columns([1.5, 4, 1.5])
        with mid_btn:
            analyze_btn = st.button("🚀 SPUSTIT ANALÝZU")
        if st.session_state.vysledky:
            if st.button("🗑 Nová analýza", use_container_width=True):
                st.session_state.vysledky = []
                st.session_state.pocet_souboru = 0
                st.session_state.obdobi_input = ''
                st.session_state.file_uploader_key += 1
                st.rerun()

    with col_main:
        if analyze_btn and uploaded_files:
            st.session_state.vysledky = []
            st.session_state.pocet_souboru = len(uploaded_files)
            st.session_state.datum_analyzy = datetime.now().strftime('%d.%m.%Y %H:%M')
            webhook_url = "https://n8n.dev.gcp.alza.cz/webhook/faktury-upload"
            with st.spinner(f"Analyzuji {len(uploaded_files)} faktur..."):
                try:
                    def get_mime(name):
                        if name.endswith('.pdf'): return "application/pdf"
                        if name.endswith('.docx'): return "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
                        if name.endswith('.xlsx'): return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        if name.endswith('.xls'): return "application/vnd.ms-excel"
                        return "application/octet-stream"
                    files = [("data", (f.name, f.getvalue(), get_mime(f.name))) for f in uploaded_files]
                    payload = {"p": st.session_state.get("obdobi_input", datetime.now().strftime('%Y-%m'))}
                    response = requests.post(webhook_url, files=files, data=payload)
                    if response.status_code == 200:
                        data = response.json()
                        st.session_state.vysledky = data if isinstance(data, list) else [data]
                        prvni = st.session_state.vysledky[0] if st.session_state.vysledky else {}
                        if prvni.get('obdobi') and str(prvni.get('obdobi')).lower() != 'n/a':
                            st.session_state.obdobi_input = prvni.get('obdobi')
                    else:
                        st.error(f"Chyba: {response.status_code}")
                except Exception as e:
                    st.error(f"Chyba spojení: {e}")
            st.rerun()

        st.subheader("📁 Digitální archiv")
        if st.session_state.vysledky:
            res = st.session_state.vysledky[0]
            if st.button("✅ ODESLAT DO TABULKY", use_container_width=False):
                if odeslat_do_google_sheets(res, sklad):
                    st.success("Uloženo do Google Sheets!")
            col_t, col_btns = st.columns([2, 1])
            with col_btns:
                col_e2, col_p2 = st.columns(2)
            df_export = pd.DataFrame(st.session_state.vysledky)
            with col_e2:
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_export.to_excel(writer, index=False, sheet_name='Energie')
                    ws = writer.sheets['Energie']
                    header_fill = PatternFill(start_color="0052CC", end_color="0052CC", fill_type="solid")
                    header_font = Font(color="FFFFFF", bold=True)
                    for col in range(1, len(df_export.columns) + 1):
                        cell = ws.cell(row=1, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                        ws.column_dimensions[get_column_letter(col)].width = 25
                    for row in range(2, len(df_export) + 2):
                        fc = "EBF3FF" if row % 2 == 0 else "FFFFFF"
                        fill = PatternFill(start_color=fc, end_color=fc, fill_type="solid")
                        for col in range(1, len(df_export.columns) + 1):
                            ws.cell(row=row, column=col).fill = fill
                periode = st.session_state.get('obdobi_input', 'export')
                st.download_button("⬇ Export Excel", data=buffer.getvalue(),
                    file_name=f"DocScan_{periode}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            with col_p2:
                pdf_buffer = io.BytesIO()
                doc = SimpleDocTemplate(pdf_buffer, pagesize=A4,
                    rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
                elements = []
                title_style = ParagraphStyle('title', fontSize=18, fontName=PDF_FONT_BOLD,
                    textColor=colors.HexColor('#0052cc'), spaceAfter=6)
                sub_style = ParagraphStyle('sub', fontSize=10, fontName=PDF_FONT,
                    textColor=colors.HexColor('#666666'), spaceAfter=20)
                elements.append(Paragraph("DocScan — Přehled energií", title_style))
                elements.append(Paragraph(
                    f"Období: {res.get('obdobi','—')}  |  Vygenerováno: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
                    sub_style))
                tisk_data = [['Kategorie', 'Parametr', 'Hodnota']]
                kats_pdf = [
                    ('Elektřina', 'el_spotreba_kwh', 'Spotřeba (kWh)'),
                    ('Elektřina', 'el_cena_sil_el_bez_dph', 'Cena sil. el. bez DPH'),
                    ('Elektřina', 'el_cena_distribuce_bez_dph', 'Cena distribuce bez DPH'),
                    ('Elektřina', 'el_cena_celkem_zaklad_kc', 'Cena celkem (Kč)'),
                    ('FSX', 'fsx_spotreba_kwh', 'Spotřeba (kWh)'),
                    ('FSX', 'fsx_cena_bez_dph', 'Cena bez DPH (Kč)'),
                    ('Plyn', 'plyn_spotreba_kwh', 'Spotřeba (kWh)'),
                    ('Plyn', 'plyn_cena_celkem_zaklad_kc', 'Cena celkem (Kč)'),
                    ('Voda', 'voda_spotreba_m3', 'Spotřeba (m³)'),
                    ('Voda', 'voda_cena_bez_dph', 'Cena bez DPH (Kč)'),
                ]
                for kat, klic, label in kats_pdf:
                    tisk_data.append([kat, label, str(res.get(klic, 'n/a'))])
                t = Table(tisk_data, colWidths=[4*cm, 8*cm, 5*cm])
                t.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0052cc')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), PDF_FONT_BOLD),
                    ('FONTSIZE', (0, 0), (-1, -1), 9),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cccccc')),
                    ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 0), (-1, -1), 6),
                    ('LEFTPADDING', (0, 0), (-1, -1), 8),
                ]))
                elements.append(t)
                elements.append(Spacer(1, 0.5*cm))
                celkem_style = ParagraphStyle('celkem', fontSize=11, fontName=PDF_FONT_BOLD,
                    textColor=colors.HexColor('#0052cc'))
                celkem_val = sum([
                    float(str(res.get(k, 0)).replace(',', '.').replace(' ', ''))
                    for k in ['el_cena_celkem_zaklad_kc', 'fsx_cena_bez_dph',
                              'plyn_cena_celkem_zaklad_kc', 'voda_cena_bez_dph']
                    if res.get(k) and str(res.get(k)).lower() != 'n/a'
                ] or [0])
                elements.append(Paragraph(
                    f"Celkem nákladů: {celkem_val:,.2f} Kč".replace(',', ' '), celkem_style))
                doc.build(elements)
                pdf_buffer.seek(0)
                st.download_button("📄 Stáhnout PDF", data=pdf_buffer.getvalue(),
                    file_name=f"DocScan_{periode}.pdf", mime="application/pdf")
            st.dataframe(df_export, use_container_width=True)
            st.write("---")
            st.subheader("📊 Finální přehled")
            cols = st.columns(4)
            kats = [
                ("⚡ Elektřina", "el_",   "el-border",    cols[0]),
                ("🏢 FSX",       "fsx_",  "fsx-border",   cols[1]),
                ("🔥 Plyn",      "plyn_", "gas-border",   cols[2]),
                ("💧 Voda",      "voda_", "water-border", cols[3]),
            ]
            for label, key, style, col in kats:
                with col:
                    st.markdown(f'<div class="energy-card {style}"><h3>{label}</h3></div>',
                        unsafe_allow_html=True)
                    for res in st.session_state.vysledky:
                        data_souboru = {k: v for k, v in res.items()
                            if k.startswith(key) and v and str(v).lower() != "n/a"}
                        if data_souboru:
                            st.markdown('<div style="margin-bottom:10px;padding:4px;">', unsafe_allow_html=True)
                            for klic, hodnota in data_souboru.items():
                                parametr = klic.replace(key, "").replace("_", " ").upper()
                                try:
                                    num = float(str(hodnota).replace(',', '.').replace(' ', ''))
                                    if 'spotreba' in klic or 'm3' in klic:
                                        hodnota_fmt = f"{num:,.0f}".replace(',', ' ')
                                    else:
                                        hodnota_fmt = f"{num:,.2f} Kč".replace(',', ' ')
                                except Exception:
                                    hodnota_fmt = str(hodnota)
                                st.markdown(
                                    f'<div style="display:flex;justify-content:space-between;'
                                    f'border-bottom:1px solid rgba(255,255,255,0.1);padding:4px 0;">'
                                    f'<span style="color:#888;font-size:0.75rem;text-transform:uppercase;">{parametr}</span>'
                                    f'<span style="color:#fff;font-weight:bold;font-size:0.85rem;">{hodnota_fmt}</span>'
                                    f'</div>', unsafe_allow_html=True)
                            st.markdown('</div>', unsafe_allow_html=True)
            st.write("")
            res = st.session_state.vysledky[0]
            text_export = f"DocScan — Výsledky analýzy\nObdobí: {res.get('obdobi','—')}\n\n"
            text_export += f"ELEKTŘINA\n  Spotřeba: {res.get('el_spotreba_kwh','n/a')} kWh\n  Cena celkem: {res.get('el_cena_celkem_zaklad_kc','n/a')} Kč\n\n"
            text_export += f"FSX\n  Spotřeba: {res.get('fsx_spotreba_kwh','n/a')} kWh\n  Cena: {res.get('fsx_cena_bez_dph','n/a')} Kč\n\n"
            text_export += f"PLYN\n  Spotřeba: {res.get('plyn_spotreba_kwh','n/a')} kWh\n  Cena celkem: {res.get('plyn_cena_celkem_zaklad_kc','n/a')} Kč\n\n"
            text_export += f"VODA\n  Spotřeba: {res.get('voda_spotreba_m3','n/a')} m³\n  Cena: {res.get('voda_cena_bez_dph','n/a')} Kč"
            st.download_button("📋 Stáhnout jako TXT", data=text_export.encode('utf-8'),
                file_name=f"DocScan_{res.get('obdobi','export')}.txt", mime="text/plain")
        else:
            st.info("Nahrajte faktury a spusťte analýzu.")

# ── FAKTURY ──────────────────────────────────────────────────────
elif st.session_state.kategorie == "Faktury":
    with col_side:
        st.markdown('<p style="color:#00c864;font-size:0.75rem;font-weight:bold;letter-spacing:2px;text-transform:uppercase;">Konfigurace</p>', unsafe_allow_html=True)
        st.file_uploader("Vložte dokumenty", accept_multiple_files=True, type=['pdf', 'docx', 'xlsx', 'xls'])
        st.markdown('<p style="color:rgba(255,255,255,0.3);font-size:0.75rem;margin-top:10px;">🔒 Dostupné po aktivaci API</p>', unsafe_allow_html=True)
    with col_main:
        st.subheader("📄 Faktury — ukázka výstupu")
        cols_f = st.columns(2)
        with cols_f[0]:
            st.markdown('<div class="energy-card fsx-border"><h4 style="color:#0084ff;">🏢 Dodavatel</h4>', unsafe_allow_html=True)
            for pole, val in [("Dodavatel", "ABC s.r.o."), ("IČ", "12345678"), ("DIČ", "CZ12345678"), ("Číslo faktury", "FAC-2026-001")]:
                st.markdown(f'<div class="preview-row"><span class="preview-label">{pole}</span><span class="preview-value">{val}</span></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with cols_f[1]:
            st.markdown('<div class="energy-card el-border"><h4 style="color:#FFD700;">💰 Platební údaje</h4>', unsafe_allow_html=True)
            for pole, val in [("Datum splatnosti", "15.02.2026"), ("Celkem bez DPH", "10 000 Kč"), ("DPH 21%", "2 100 Kč"), ("Celkem s DPH", "12 100 Kč"), ("Číslo účtu", "123456789/0800"), ("Variabilní symbol", "20260001")]:
                st.markdown(f'<div class="preview-row"><span class="preview-label">{pole}</span><span class="preview-value">{val}</span></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        st.info("⏳ Funkce bude aktivní po připojení Anthropic API.")

# ── SMLOUVY ──────────────────────────────────────────────────────
elif st.session_state.kategorie == "Smlouvy":
    with col_side:
        st.markdown('<p style="color:#00c864;font-size:0.75rem;font-weight:bold;letter-spacing:2px;text-transform:uppercase;">Konfigurace</p>', unsafe_allow_html=True)
        st.file_uploader("Vložte dokumenty", accept_multiple_files=True, type=['pdf', 'docx', 'xlsx', 'xls'])
        st.markdown('<p style="color:rgba(255,255,255,0.3);font-size:0.75rem;margin-top:10px;">🔒 Dostupné po aktivaci API</p>', unsafe_allow_html=True)
    with col_main:
        st.subheader("📋 Smlouvy — ukázka výstupu")
        cols_s = st.columns(2)
        with cols_s[0]:
            st.markdown('<div class="energy-card gas-border"><h4 style="color:#FF5722;">📝 Smluvní strany</h4>', unsafe_allow_html=True)
            for pole, val in [("Objednatel", "XYZ a.s."), ("Zhotovitel", "ABC s.r.o."), ("Datum podpisu", "01.01.2026"), ("Platnost do", "31.12.2026")]:
                st.markdown(f'<div class="preview-row"><span class="preview-label">{pole}</span><span class="preview-value">{val}</span></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        with cols_s[1]:
            st.markdown('<div class="energy-card water-border"><h4 style="color:#00BFFF;">📌 Klíčové podmínky</h4>', unsafe_allow_html=True)
            for pole, val in [("Předmět", "Dodávka služeb"), ("Hodnota", "120 000 Kč/rok"), ("Výpovědní lhůta", "3 měsíce"), ("Obnova", "Automatická")]:
                st.markdown(f'<div class="preview-row"><span class="preview-label">{pole}</span><span class="preview-value">{val}</span></div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)
        st.info("⏳ Funkce bude aktivní po připojení Anthropic API.")

# ── OOPP & MCDP ──────────────────────────────────────────────────
elif st.session_state.kategorie == "OOPP & MČDP":
    with col_side:
        st.markdown('<p style="color:#00c864;font-size:0.75rem;font-weight:bold;letter-spacing:2px;text-transform:uppercase;">Konfigurace</p>', unsafe_allow_html=True)
        sklad_oopp = st.selectbox("Sklad:", ["CZLC4", "LCÚ", "LCZ", "SKLC3"], key="sklad_oopp")
        rezim = st.radio("Režim:", ["Výdej MČDP", "Evidence OOPP", "Tisk protokolu MČDP", "Tisk protokolu OOPP"])

    with col_main:
        PODPIS_URL = "https://janasiva4.github.io/DocScan-Alza/podpis_2fa.html"

        # ═══════════════ VÝDEJ MČDP ═══════════════
        if rezim == "Výdej MČDP":
            st.subheader("🧴 Výdej MČDP — kvartální")
            if 'mcdp_reset' not in st.session_state:
                st.session_state.mcdp_reset = 0
            zamestnanec = st.text_input("Zaměstnanec (jméno a příjmení)", key=f"zam_{st.session_state.mcdp_reset}", autocomplete="off")
            email_zam   = st.text_input("Email zaměstnance", placeholder="jan.novak@firma.cz", key=f"email_{st.session_state.mcdp_reset}", autocomplete="off")
            stredisko   = st.text_input("Středisko", placeholder="např. Sklad A — příjem", key=f"stredisko_{st.session_state.mcdp_reset}", autocomplete="off")
            user        = st.text_input("Uživatel / osobní číslo", placeholder="např. 12345", key=f"user_{st.session_state.mcdp_reset}", autocomplete="off")
            rok_akt = datetime.now().year
            kvartal_sel = st.selectbox("Kvartál", [
                f"Q1 / {rok_akt}", f"Q2 / {rok_akt}", f"Q3 / {rok_akt}", f"Q4 / {rok_akt}",
                f"Q1 / {rok_akt+1}", f"Q2 / {rok_akt+1}", f"Q3 / {rok_akt+1}", f"Q4 / {rok_akt+1}"])
            st.write("**Vydávané položky:**")
            c1, c2 = st.columns(2)
            rucnik  = c1.checkbox("1x Ručník Siguro 50x100cm", value=True)
            mydlo   = c2.checkbox("1x Tekuté mýdlo", value=True)
            ariel   = c1.checkbox("1x Ariel tablety 60 ks", value=True)
            krem    = c2.checkbox("1x Krém Indulona", value=True)
            solvina = c1.checkbox("1x Abrazivní pasta Solvina", value=True)
            # Kolonka velikost pro ručník
            velikost_rucnik = st.text_input("Velikost ručníku (volitelně, pro tisk)",
                placeholder="např. 50x100 cm", key=f"velrucnik_{st.session_state.mcdp_reset}")
            vedouci = st.text_input("Zadal / vedoucí")

            if zamestnanec and email_zam:
                polozky_list = []
                if rucnik:  polozky_list.append("Ručník Siguro")
                if mydlo:   polozky_list.append("Tekuté mýdlo")
                if ariel:   polozky_list.append("Ariel 60 ks")
                if krem:    polozky_list.append("Krém Indulona")
                if solvina: polozky_list.append("Solvina")
                qr_data = {"jmeno": zamestnanec, "email": email_zam, "stredisko": stredisko,
                           "user": user, "sklad": sklad_oopp, "kvartal": kvartal_sel,
                           "polozky": ", ".join(polozky_list)}
                qr_json = json.dumps(qr_data, ensure_ascii=False)
                qr_payload = base64.b64encode(qr_json.encode('utf-8')).decode('ascii')
                qr_url = f"{PODPIS_URL}?d={qr_payload}"
                qr = qrcode.QRCode(version=1, box_size=6, border=2)
                qr.add_data(qr_url)
                qr.make(fit=True)
                qr_img = qr.make_image(fill_color="#1a3a6b", back_color="white")
                buf_qr = io.BytesIO()
                qr_img.save(buf_qr, format="PNG")
                st.write("---")
                col_qr, col_info = st.columns([1, 2])
                with col_qr:
                    st.image(buf_qr.getvalue(), width=180, caption="Zaměstnanec naskenuje pro podpis")
                with col_info:
                    polozky_str = ", ".join(polozky_list)
                    st.markdown(
                        f'<div class="energy-card oopp-border" style="padding:12px;">'
                        f'<p style="color:#00c864;font-size:0.75rem;font-weight:bold;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">Čeká na 2FA podpis</p>'
                        f'<p style="color:#fff;font-size:0.9rem;"><b>{zamestnanec}</b></p>'
                        f'<p style="color:#aaa;font-size:0.8rem;">{email_zam}</p>'
                        f'<p style="color:#aaa;font-size:0.8rem;margin-top:4px;">{kvartal_sel} · {sklad_oopp}</p>'
                        f'<p style="color:#aaa;font-size:0.75rem;margin-top:4px;">{polozky_str}</p>'
                        f'</div>', unsafe_allow_html=True)
                st.write("---")

            col_btn1, col_btn2 = st.columns(2)
            with col_btn1:
                if st.button("✅ ODESLAT DO GOOGLE SHEETS", use_container_width=True):
                    if not zamestnanec:
                        st.warning("Zadej jméno zaměstnance.")
                    else:
                        data = {"zamestnanec": zamestnanec, "email": email_zam, "kvartal": kvartal_sel,
                                "rucnik": rucnik, "mydlo": mydlo, "ariel": ariel, "krem": krem,
                                "solvina": solvina, "podpis": True, "zadal": vedouci}
                        if odeslat_mcdp_do_sheets(data, sklad_oopp):
                            st.success(f"✅ Záznam uložen — {zamestnanec} · {kvartal_sel}")
                            st.session_state.mcdp_reset += 1
                            st.rerun()
            with col_btn2:
                if zamestnanec:
                    pdf_bytes = generovat_pdf_protokol(
                        zamestnanec=zamestnanec, sklad=sklad_oopp, kvartal=kvartal_sel,
                        vydane_polozky={"rucnik": rucnik, "mydlo": mydlo, "ariel": ariel,
                                        "krem": krem, "solvina": solvina},
                        vedouci=vedouci,
                        velikosti={"rucnik": velikost_rucnik} if velikost_rucnik else None)
                    jmeno_souboru = zamestnanec.replace(" ", "_")
                    st.download_button(
                        label="📄 Stáhnout PDF protokol",
                        data=bytes(pdf_bytes),
                        file_name=f"Protokol_MCDP_{jmeno_souboru}_{kvartal_sel[:2]}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key=f"dl_mcdp_{st.session_state.mcdp_reset}")

      # ═══════════════ EVIDENCE OOPP ═══════════════
        elif rezim == "Evidence OOPP":
            st.subheader("🦺 Evidence OOPP — výdej pomůcek")
            if 'oopp_reset' not in st.session_state:
                st.session_state.oopp_reset = 0
            zamestnanec2 = st.text_input("Zaměstnanec (jméno a příjmení)", key=f"zam2_{st.session_state.oopp_reset}", autocomplete="off")
            email_zam2   = st.text_input("Email zaměstnance", placeholder="jan.novak@firma.cz", key=f"email2_{st.session_state.oopp_reset}", autocomplete="off")
            stredisko2   = st.text_input("Středisko", placeholder="např. Sklad A — příjem", key=f"stredisko2_{st.session_state.oopp_reset}", autocomplete="off")
            user2        = st.text_input("Uživatel / osobní číslo", placeholder="např. 12345", key=f"user2_{st.session_state.oopp_reset}", autocomplete="off")
            vedouci2     = st.text_input("Zadal / vedoucí", key=f"vedouci2_{st.session_state.oopp_reset}", autocomplete="off")
            st.write("**Vydávané pomůcky:**")
            pomucky_def = [
                ("Oděv pracovní (montérky)", "odev", None),
                ("Rukavice bezpečnostní", "rukavice", None),
                ("Kabát proti chladu", "kabat", 24),
                ("Tričko", "tricko", 12),
                ("Mikina", "mikina", 6),
                ("Čepice / kšiltovka", "cepice", 24),
                ("Ochranné brýle", "bryle", None),
                ("Kraťasy", "kratasy", 12),
                ("Thermo", "thermo", 12),
                ("Bezpečnostní obuv", "obuv", 12),
            ]
            o1, o2 = st.columns(2)
            vydane = {}
            velikosti_vyd = {}
            for i, (nazev, klic, exp_mes) in enumerate(pomucky_def):
                col = o1 if i % 2 == 0 else o2
                if exp_mes and exp_mes >= 12:
                    exp_info = f" ({exp_mes//12}r)"
                elif exp_mes:
                    exp_info = f" ({exp_mes}m)"
                else:
                    exp_info = " (dle potřeby)"
                vydane[klic] = col.checkbox(f"{nazev}{exp_info}", key=f"oopp_{klic}_{st.session_state.oopp_reset}")
                if vydane[klic]:
                    velikosti_vyd[klic] = col.text_input(f"Velikost — {nazev}",
                        key=f"vel_{klic}_{st.session_state.oopp_reset}",
                        placeholder="např. L, XL, 42, …")

            def exp_datum(mesice):
                if not mesice:
                    return None
                d = date.today()
                mes = d.month + mesice
                rok_exp = d.year + (mes - 1) // 12
                mes_exp = (mes - 1) % 12 + 1
                return f"{mes_exp:02d}/{rok_exp}"

            # QR kód se zobrazí pouze pokud je vyplněno jméno a email
            if zamestnanec2 and email_zam2:
                vydane_nazvy = [nazev for nazev, klic, _ in pomucky_def if vydane.get(klic)]
                oopp_qr_data = {"jmeno": zamestnanec2, "email": email_zam2, "stredisko": stredisko2,
                                "user": user2, "sklad": sklad_oopp,
                                "kvartal": f"OOPP {datetime.now().strftime('%m/%Y')}",
                                "polozky": ", ".join(vydane_nazvy) if vydane_nazvy else "—"}
                qr_json2 = json.dumps(oopp_qr_data, ensure_ascii=False)
                qr_payload2 = base64.b64encode(qr_json2.encode('utf-8')).decode('ascii')
                qr_url2 = f"{PODPIS_URL}?d={qr_payload2}"
                qr2 = qrcode.QRCode(version=1, box_size=6, border=2)
                qr2.add_data(qr_url2)
                qr2.make(fit=True)
                qr_img2 = qr2.make_image(fill_color="#0D4F1C", back_color="white")
                buf_qr2 = io.BytesIO()
                qr_img2.save(buf_qr2, format="PNG")
                st.write("---")
                col_qr2, col_info2 = st.columns([1, 2])
                with col_qr2:
                    st.image(buf_qr2.getvalue(), width=180, caption="Zaměstnanec naskenuje pro podpis")
                with col_info2:
                    nazvy_str = ", ".join(vydane_nazvy) if vydane_nazvy else "—"
                    st.markdown(
                        f'<div class="energy-card oopp-border" style="padding:12px;">'
                        f'<p style="color:#00c864;font-size:0.75rem;font-weight:bold;text-transform:uppercase;letter-spacing:1px;margin-bottom:8px;">Čeká na 2FA podpis</p>'
                        f'<p style="color:#fff;font-size:0.9rem;"><b>{zamestnanec2}</b></p>'
                        f'<p style="color:#aaa;font-size:0.8rem;">{email_zam2}</p>'
                        f'<p style="color:#aaa;font-size:0.75rem;margin-top:4px;">{nazvy_str}</p>'
                        f'</div>', unsafe_allow_html=True)
                st.write("---")

            # === TLAČÍTKA JSOU VŽDY VIDITELNÁ (mimo if blok) — fungují stejně jako v MČDP ===
            col_btn_o1, col_btn_o2 = st.columns(2)
            with col_btn_o1:
                if st.button("✅ ULOŽIT DO EVIDENCE", use_container_width=True):
                    # Validace vstupů — stejně jako u MČDP
                    if not zamestnanec2:
                        st.warning("Zadej jméno zaměstnance.")
                    elif not email_zam2:
                        st.warning("Zadej email zaměstnance.")
                    elif not any(vydane.get(klic) for _, klic, _ in pomucky_def):
                        st.warning("Označ alespoň jednu pomůcku k vydání.")
                    else:
                        ulozeno = 0
                        chyby = 0
                        for nazev, klic, exp_mes in pomucky_def:
                            if vydane.get(klic):
                                exp = exp_datum(exp_mes)
                                data_oopp = {
                                    "zamestnanec": zamestnanec2,
                                    "email": email_zam2,
                                    "stredisko": stredisko2,
                                    "user": user2,
                                    "pomucka": nazev,
                                    "velikost": velikosti_vyd.get(klic, ""),
                                    "expirace": exp or "",
                                    "podpis": True,
                                    "zadal": vedouci2,
                                }
                                if odeslat_oopp_do_sheets(data_oopp, sklad_oopp):
                                    ulozeno += 1
                                else:
                                    chyby += 1
                        if ulozeno > 0 and chyby == 0:
                            st.success(f"✅ Uloženo {ulozeno} pomůcek do Google Sheets — {zamestnanec2}")
                            st.session_state.oopp_reset += 1
                            st.rerun()
                        elif ulozeno > 0 and chyby > 0:
                            st.warning(f"Uloženo {ulozeno} pomůcek, {chyby} se nepodařilo odeslat.")
                        else:
                            st.error("Záznam se nepodařilo odeslat do Google Sheets.")
            with col_btn_o2:
                if zamestnanec2:
                    expirace_dict = {klic: exp_datum(exp_mes) or ""
                                     for nazev, klic, exp_mes in pomucky_def if vydane.get(klic)}
                    pdf_oopp_bytes = generovat_pdf_oopp(
                        zamestnanec=zamestnanec2, email=email_zam2, sklad=sklad_oopp,
                        vydane_pomucky=vydane, velikosti_oopp=velikosti_vyd,
                        expirace_oopp=expirace_dict, vedouci=vedouci2,
                        stredisko=stredisko2, osobni_cislo=user2)
                    jmeno_soub_o = zamestnanec2.replace(" ", "_")
                    st.download_button(
                        label="📄 Stáhnout PDF protokol OOPP",
                        data=bytes(pdf_oopp_bytes),
                        file_name=f"Protokol_OOPP_{jmeno_soub_o}.pdf",
                        mime="application/pdf",
                        use_container_width=True,
                        key=f"dl_oopp_{st.session_state.oopp_reset}")
            else:
                st.info("Vyplň jméno a email zaměstnance pro zobrazení QR kódu.")

        # ═══════════════ TISK PROTOKOLU MČDP ═══════════════
        elif rezim == "Tisk protokolu MČDP":
            st.subheader("🖨️ Generátor předávacího protokolu — MČDP")
            st.markdown('<p style="color:rgba(255,255,255,0.5);font-size:0.85rem;">Vyplň údaje — dostaneš PDF připravené k tisku a podpisu zaměstnance.</p>', unsafe_allow_html=True)
            zam_tisk = st.text_input("Zaměstnanec")
            rok_akt2 = datetime.now().year
            kv_tisk = st.selectbox("Kvartál", [
                f"Q1 / {rok_akt2}", f"Q2 / {rok_akt2}", f"Q3 / {rok_akt2}", f"Q4 / {rok_akt2}",
                f"Q1 / {rok_akt2+1}", f"Q2 / {rok_akt2+1}", f"Q3 / {rok_akt2+1}", f"Q4 / {rok_akt2+1}"],
                key="kv_tisk")
            ved_tisk = st.text_input("Vedoucí", key="ved_tisk")
            st.write("**Položky pro protokol:**")
            t1, t2 = st.columns(2)
            cb1 = t1.checkbox("Ručník Siguro", value=True, key="p1")
            cb2 = t2.checkbox("Tekuté mýdlo",  value=True, key="p2")
            cb3 = t1.checkbox("Ariel 60 ks",   value=True, key="p3")
            cb4 = t2.checkbox("Krém Indulona",  value=True, key="p4")
            cb5 = t1.checkbox("Solvina",        value=True, key="p5")
            vel_rucnik_tisk = st.text_input("Velikost ručníku (volitelně)",
                placeholder="např. 50×100 cm", key="vel_rucnik_tisk")

            pdf_tisk = generovat_pdf_protokol(
                zamestnanec=zam_tisk or "—",
                sklad=sklad_oopp,
                kvartal=kv_tisk,
                vydane_polozky={"rucnik": cb1, "mydlo": cb2, "ariel": cb3, "krem": cb4, "solvina": cb5},
                vedouci=ved_tisk,
                velikosti={"rucnik": vel_rucnik_tisk} if vel_rucnik_tisk else None
            )
            jmeno_souboru = (zam_tisk or "protokol").replace(" ", "_")
            st.download_button(
                label="📄 Stáhnout PDF protokol k tisku",
                data=bytes(pdf_tisk) if pdf_tisk else b"",
                file_name=f"Protokol_MCDP_{jmeno_souboru}.pdf",
                mime="application/pdf",
                use_container_width=False,
                disabled=not zam_tisk,
                key=f"dl_tisk_{zam_tisk or 'empty'}"
            )
            if not zam_tisk:
                st.info("Zadej jméno zaměstnance pro aktivaci tlačítka stažení.")

        # ═══════════════ TISK PROTOKOLU OOPP ═══════════════
        elif rezim == "Tisk protokolu OOPP":
            st.subheader("🖨️ Generátor předávacího protokolu — OOPP")
            st.markdown('<p style="color:rgba(255,255,255,0.5);font-size:0.85rem;">Vyplň údaje — dostaneš PDF připravené k tisku a podpisu zaměstnance. Velikosti lze doplnit ručně na papíru.</p>', unsafe_allow_html=True)
            zam_oopp_tisk = st.text_input("Zaměstnanec", key="zam_oopp_tisk")
            email_oopp_tisk = st.text_input("Email", key="email_oopp_tisk", placeholder="jan.novak@firma.cz")
            stredisko_oopp_tisk = st.text_input("Středisko", key="stredisko_oopp_tisk")
            user_oopp_tisk = st.text_input("Osobní číslo", key="user_oopp_tisk")
            ved_oopp_tisk = st.text_input("Vedoucí", key="ved_oopp_tisk")

            pomucky_tisk_def = [
                ("Oděv pracovní (montérky)", "odev", None),
                ("Rukavice bezpečnostní", "rukavice", None),
                ("Kabát proti chladu", "kabat", 24),
                ("Tričko", "tricko", 12),
                ("Mikina", "mikina", 6),
                ("Čepice / kšiltovka", "cepice", 24),
                ("Ochranné brýle", "bryle", None),
                ("Kraťasy", "kratasy", 12),
                ("Thermo prádlo", "thermo", 12),
                ("Bezpečnostní obuv", "obuv", 12),
            ]
            st.write("**Pomůcky pro protokol:**")
            c_o1, c_o2 = st.columns(2)
            vydane_tisk = {}
            velikosti_tisk = {}
            for i, (nazev, klic, _) in enumerate(pomucky_tisk_def):
                col = c_o1 if i % 2 == 0 else c_o2
                vydane_tisk[klic] = col.checkbox(nazev, key=f"tiskoopp_{klic}", value=True)
                if vydane_tisk[klic]:
                    velikosti_tisk[klic] = col.text_input(f"Velikost — {nazev}",
                        key=f"tiskvel_{klic}",
                        placeholder="např. L, 42, …")

            def exp_datum_tisk(mesice):
                if not mesice:
                    return ""
                d = date.today()
                mes = d.month + mesice
                rok_exp = d.year + (mes - 1) // 12
                mes_exp = (mes - 1) % 12 + 1
                return f"{mes_exp:02d}/{rok_exp}"

            expirace_tisk = {klic: exp_datum_tisk(exp_mes)
                             for nazev, klic, exp_mes in pomucky_tisk_def if vydane_tisk.get(klic)}

            pdf_oopp_tisk = generovat_pdf_oopp(
                zamestnanec=zam_oopp_tisk or "—",
                email=email_oopp_tisk,
                sklad=sklad_oopp,
                vydane_pomucky=vydane_tisk,
                velikosti_oopp=velikosti_tisk,
                expirace_oopp=expirace_tisk,
                vedouci=ved_oopp_tisk,
                stredisko=stredisko_oopp_tisk,
                osobni_cislo=user_oopp_tisk
            )
            jmeno_soub_tisk = (zam_oopp_tisk or "protokol").replace(" ", "_")
            st.download_button(
                label="📄 Stáhnout PDF protokol OOPP k tisku",
                data=bytes(pdf_oopp_tisk) if pdf_oopp_tisk else b"",
                file_name=f"Protokol_OOPP_{jmeno_soub_tisk}.pdf",
                mime="application/pdf",
                use_container_width=False,
                disabled=not zam_oopp_tisk,
                key=f"dl_tisk_oopp_{zam_oopp_tisk or 'empty'}"
            )
            if not zam_oopp_tisk:
                st.info("Zadej jméno zaměstnance pro aktivaci tlačítka stažení.")
